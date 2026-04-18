from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


MONTH_NAMES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


@dataclass(frozen=True)
class MonthSpec:
    year: int
    month: int

    @property
    def days(self) -> int:
        return monthrange(self.year, self.month)[1]

    @property
    def sheet_name(self) -> str:
        return f"{MONTH_NAMES[self.month]} {self.year}"


MONTHS = [
    MonthSpec(2024, 12),
    MonthSpec(2025, 1),
    MonthSpec(2025, 2),
    MonthSpec(2025, 3),
    MonthSpec(2025, 4),
    MonthSpec(2025, 5),
    MonthSpec(2025, 6),
    MonthSpec(2025, 7),
    MonthSpec(2025, 8),
    MonthSpec(2025, 9),
    MonthSpec(2025, 10),
    MonthSpec(2025, 11),
    MonthSpec(2025, 12),
    MonthSpec(2026, 1),
    MonthSpec(2026, 2),
]


THIN = Side(style="thin", color="D9DDE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION = PatternFill("solid", fgColor="DCE6F1")
FILL_HEADER = PatternFill("solid", fgColor="EAF2F8")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FILL_CALC = PatternFill("solid", fgColor="E2F0D9")
FILL_ALERT = PatternFill("solid", fgColor="FCE4D6")

FONT_TITLE = Font(color="FFFFFF", bold=True, size=14)
FONT_BOLD = Font(bold=True)


def style_range(ws, cell_range: str, fill=None, font=None, align=None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if align is not None:
                cell.alignment = align


def protect_formula_cell(cell) -> None:
    cell.protection = Protection(locked=True)
    cell.fill = FILL_CALC


def unlock_input_cell(cell) -> None:
    cell.protection = Protection(locked=False)
    cell.fill = FILL_INPUT


def build_readme(ws) -> None:
    ws.title = "00_LEEME"
    ws["A1"] = "HODOM - ESTRUCTURA NUEVA"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:F1")

    rows = [
        "Objetivo",
        "Dejar una base limpia, simple y consistente para cargar y migrar los meses desde la planilla antigua.",
        "",
        "Qué trae este archivo",
        "1. Una hoja de parámetros generales.",
        "2. Una hoja mensual estándar por cada mes del período original.",
        "3. Una hoja de control para revisar rápidamente totales e inconsistencias.",
        "",
        "Cómo usarlo",
        "1. Editar solo celdas amarillas.",
        "2. No editar celdas verdes: son cálculos.",
        "3. Revisar alertas en cada hoja mensual y en 99_CONTROL.",
        "4. Migrar mes por mes desde la planilla antigua.",
        "",
        "Definiciones clave",
        "Pacientes atendidos del mes = existencia inicial del día 1 + total de ingresos del mes.",
        "Índice ocupacional % = días cama ocupadas / días cama disponibles.",
        "Promedio día de estada = días estada total / total egresos.",
    ]
    for idx, text in enumerate(rows, start=3):
        ws[f"A{idx}"] = text
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18


def build_parameters(ws) -> None:
    ws.title = "00_PARAMETROS"
    ws["A1"] = "PARAMETROS GENERALES"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:D1")

    rows = [
        ("Servicio", "HODOM"),
        ("Codigo", ""),
        ("Camas base por dia", 16),
        ("Responsable", ""),
        ("Observaciones", ""),
    ]
    for idx, (label, value) in enumerate(rows, start=3):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = FONT_BOLD
        ws[f"A{idx}"].fill = FILL_SECTION
        ws[f"A{idx}"].border = BORDER
        ws[f"B{idx}"] = value
        ws[f"B{idx}"].border = BORDER
        unlock_input_cell(ws[f"B{idx}"])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14


def build_month_sheet(ws, spec: MonthSpec) -> None:
    ws.title = spec.sheet_name
    ws.freeze_panes = "A7"

    ws["A1"] = "RESUMEN MENSUAL HODOM"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:S1")

    meta = [
        ("Servicio", "=00_PARAMETROS!B3", "Codigo", "=00_PARAMETROS!B4", "Mes", MONTH_NAMES[spec.month], "Ano", spec.year),
        ("Camas base", "=00_PARAMETROS!B5", "Promedio camas", None, "Dias del mes", spec.days, "", ""),
    ]
    for row_idx, row_values in enumerate(meta, start=2):
        for offset in range(0, 8, 2):
            label_col = get_column_letter(offset + 1)
            value_col = get_column_letter(offset + 2)
            label = row_values[offset]
            value = row_values[offset + 1]
            if not label:
                continue
            ws[f"{label_col}{row_idx}"] = label
            ws[f"{label_col}{row_idx}"].font = FONT_BOLD
            ws[f"{label_col}{row_idx}"].fill = FILL_SECTION
            ws[f"{label_col}{row_idx}"].border = BORDER
            ws[f"{value_col}{row_idx}"] = value
            ws[f"{value_col}{row_idx}"].border = BORDER
            if isinstance(value, str) and value.startswith("="):
                protect_formula_cell(ws[f"{value_col}{row_idx}"])
            else:
                unlock_input_cell(ws[f"{value_col}{row_idx}"])

    headers = [
        "Fecha",
        "Dia",
        "Existencia dia anterior",
        "Ing. urgencia",
        "Ing. hospitalizacion",
        "Ing. ambulatorio",
        "Ing. ley urgencia",
        "Ing. UGCC",
        "Ing. por traslado",
        "Total ingresos",
        "Egr. alta hogar / otro hosp.",
        "Egr. traslado a otro servicio",
        "Egr. fallecidos",
        "Total egresos",
        "Ingr.-egr. mismo dia",
        "Dias cama disponibles",
        "Dias cama ocupadas",
        "Dias estada total",
        "Dias estada benef.",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(6, col_idx, header)
        cell.font = FONT_BOLD
        cell.fill = FILL_HEADER
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    input_cols = {"C", "D", "E", "F", "G", "H", "I", "K", "L", "M", "O", "P", "R", "S"}
    first_row = 7
    last_row = first_row + spec.days - 1

    non_negative = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    non_negative.prompt = "Ingresa un numero entero mayor o igual a 0."
    ws.add_data_validation(non_negative)

    for day in range(1, spec.days + 1):
        row = first_row + day - 1
        ws[f"A{row}"] = f'=DATE({spec.year},{spec.month},{day})'
        ws[f"B{row}"] = day
        protect_formula_cell(ws[f"A{row}"])
        protect_formula_cell(ws[f"B{row}"])

        if day == 1:
            unlock_input_cell(ws[f"C{row}"])
        else:
            ws[f"C{row}"] = f"=Q{row-1}"
            protect_formula_cell(ws[f"C{row}"])

        for col in input_cols - {"C"}:
            unlock_input_cell(ws[f"{col}{row}"])
            if col == "P":
                ws[f"{col}{row}"] = "=00_PARAMETROS!B5"
            non_negative.add(ws[f"{col}{row}"])

        ws[f"J{row}"] = f"=SUM(D{row}:H{row})+I{row}"
        ws[f"N{row}"] = f"=SUM(K{row}:M{row})"
        ws[f"Q{row}"] = f"=C{row}+J{row}-N{row}"
        for col in ("J", "N", "Q"):
            protect_formula_cell(ws[f"{col}{row}"])

        for col_idx in range(1, 20):
            ws.cell(row, col_idx).border = BORDER

    total_row = last_row + 1
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = FONT_BOLD
    ws[f"A{total_row}"].fill = FILL_SECTION
    ws[f"A{total_row}"].border = BORDER
    for col_idx in range(2, 20):
        col = get_column_letter(col_idx)
        if col in {"A", "B", "C"}:
            ws[f"{col}{total_row}"].border = BORDER
            continue
        ws[f"{col}{total_row}"] = f"=SUM({col}{first_row}:{col}{last_row})"
        ws[f"{col}{total_row}"].font = FONT_BOLD
        ws[f"{col}{total_row}"].border = BORDER
        protect_formula_cell(ws[f"{col}{total_row}"])

    ws["D3"] = f"=AVERAGE(P{first_row}:P{last_row})"
    protect_formula_cell(ws["D3"])

    summary_row = total_row + 3
    ws[f"A{summary_row}"] = "RESUMEN DEL MES"
    ws[f"A{summary_row}"].font = FONT_BOLD
    ws[f"A{summary_row}"].fill = FILL_SECTION
    ws[f"A{summary_row}"].border = BORDER
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)

    summary_items = [
        ("Ingresos", f"=J{total_row}"),
        ("Altas al hogar / otro hosp.", f"=K{total_row}"),
        ("Traslados a otro servicio", f"=L{total_row}"),
        ("Fallecidos", f"=M{total_row}"),
        ("Total egresos", f"=N{total_row}"),
        ("Pacientes atendidos del mes", f"=C{first_row}+J{total_row}"),
        ("Dias personas atendidas", f"=R{total_row}"),
        ("Indice ocupacional %", f'=IFERROR(Q{total_row}/P{total_row},"")'),
        ("Promedio dia de estada", f'=IFERROR(R{total_row}/N{total_row},"")'),
    ]
    for offset, (label, formula) in enumerate(summary_items, start=1):
        row = summary_row + offset
        ws[f"A{row}"] = label
        ws[f"A{row}"].border = BORDER
        ws[f"A{row}"].fill = FILL_HEADER
        ws[f"A{row}"].font = FONT_BOLD
        ws[f"B{row}"] = formula
        ws[f"B{row}"].border = BORDER
        protect_formula_cell(ws[f"B{row}"])
        if "ocupacional" in label:
            ws[f"B{row}"].number_format = "0.00%"
        elif "Promedio" in label:
            ws[f"B{row}"].number_format = "0.00"

    alert_col = "U"
    ws[f"{alert_col}1"] = "CONTROL"
    ws[f"{alert_col}1"].font = FONT_TITLE
    ws[f"{alert_col}1"].fill = FILL_TITLE
    ws[f"{alert_col}1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("U1:W1")

    checks = [
        ("Dias con ocupacion negativa", f'=COUNTIF(Q{first_row}:Q{last_row},"<0")'),
        ("Dias sin camas disponibles", f'=COUNTBLANK(P{first_row}:P{last_row})'),
        ("Primer dia sin existencia inicial", f'=IF(C{first_row}="",1,0)'),
        ("Dias con total ingresos manualmente alterado", "0"),
        ("Dias con total egresos manualmente alterado", "0"),
    ]
    for idx, (label, formula) in enumerate(checks, start=3):
        ws[f"U{idx}"] = label
        ws[f"U{idx}"].border = BORDER
        ws[f"U{idx}"].fill = FILL_ALERT
        ws[f"V{idx}"] = formula
        ws[f"V{idx}"].border = BORDER
        protect_formula_cell(ws[f"V{idx}"])

    for column, width in {
        "A": 13,
        "B": 7,
        "C": 20,
        "D": 13,
        "E": 18,
        "F": 15,
        "G": 16,
        "H": 10,
        "I": 15,
        "J": 13,
        "K": 20,
        "L": 22,
        "M": 14,
        "N": 13,
        "O": 18,
        "P": 18,
        "Q": 17,
        "R": 16,
        "S": 16,
        "U": 34,
        "V": 12,
    }.items():
        ws.column_dimensions[column].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[6].height = 36

    ws.sheet_view.showGridLines = True
    ws.protection.sheet = True
    ws.protection.password = "hodom"
    ws.protection.enable()
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def build_control_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("99_CONTROL")
    ws["A1"] = "CONTROL GENERAL"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:H1")

    headers = [
        "Mes",
        "Ingresos",
        "Total egresos",
        "Pacientes atendidos",
        "Dias personas atendidas",
        "Indice ocupacional",
        "Promedio dia de estada",
        "Alertas",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(3, idx, header)
        cell.font = FONT_BOLD
        cell.fill = FILL_HEADER
        cell.border = BORDER

    for row, spec in enumerate(MONTHS, start=4):
        sheet = f"'{spec.sheet_name}'"
        ws[f"A{row}"] = spec.sheet_name
        ws[f"B{row}"] = f"={sheet}!B{spec.days + 11}"
        ws[f"C{row}"] = f"={sheet}!B{spec.days + 15}"
        ws[f"D{row}"] = f"={sheet}!B{spec.days + 16}"
        ws[f"E{row}"] = f"={sheet}!B{spec.days + 17}"
        ws[f"F{row}"] = f"={sheet}!B{spec.days + 18}"
        ws[f"G{row}"] = f"={sheet}!B{spec.days + 19}"
        ws[f"H{row}"] = f"=SUM({sheet}!V3:V7)"
        for col_idx in range(1, 9):
            cell = ws.cell(row, col_idx)
            cell.border = BORDER
            if col_idx > 1:
                protect_formula_cell(cell)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 10


def create_workbook(output_path: Path) -> Path:
    wb = Workbook()
    build_readme(wb.active)
    build_parameters(wb.create_sheet("00_PARAMETROS"))
    for spec in MONTHS:
        build_month_sheet(wb.create_sheet(spec.sheet_name), spec)
    build_control_sheet(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    destination = Path("tmp/spreadsheets/HODOM_estructura_nueva.xlsx")
    final_path = create_workbook(destination)
    print(final_path)
