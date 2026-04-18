from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from generate_hodom_ideal_workbook import MONTHS, MonthSpec


SOURCE_PATH = Path("/Users/felixsanhueza/Downloads/HOMOM 2026 ACT.xlsx")
TEMPLATE_PATH = Path("/Users/felixsanhueza/Developer/kora/tmp/spreadsheets/HODOM_ideal.xlsx")
OUTPUT_DIR = Path("/Users/felixsanhueza/Developer/kora/tmp/spreadsheets")
OUTPUT_WORKBOOK = OUTPUT_DIR / "HODOM_ideal_migrado.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "HODOM_migration_report.csv"
OUTPUT_MD = OUTPUT_DIR / "HODOM_migration_report.md"

KNOWN_MANUAL_SUMMARY_MONTHS = {
    "ABRIL 2025",
    "MAYO 2025",
    "JULIO 2025",
    "OCTUBRE 2025",
    "NOVIEMBRE 2025",
    "DICIEMBRE 2025",
    "ENERO 2026",
    "FEBRERO 2026",
}

SOURCE_STATUS_COLUMN_MAP = {
    "urgencia": 3,
    "hospitalizacion": 4,
    "ambulatorio": 5,
    "ley_urgencia": 6,
    "ugcc": 7,
    "procedencia_total": 8,
    "traslado_ingreso": 9,
    "total_ingresos": 10,
    "alta_hogar": 11,
    "traslado_servicio": 12,
    "fallecidos": 13,
    "total_egresos": 14,
    "mismo_dia": 15,
    "camas_disponibles": 16,
    "camas_ocupadas": 17,
    "dias_estada_total": 18,
    "dias_estada_benef": 19,
}

TARGET_INPUT_COLUMN_MAP = {
    "urgencia": "D",
    "hospitalizacion": "E",
    "ambulatorio": "F",
    "ley_urgencia": "G",
    "ugcc": "H",
    "traslado_ingreso": "I",
    "total_ingresos_informado": "K",
    "alta_hogar": "L",
    "traslado_servicio": "M",
    "fallecidos": "N",
    "total_egresos_informado": "P",
    "mismo_dia": "Q",
    "camas_disponibles": "R",
    "dias_estada_total": "T",
    "dias_estada_benef": "U",
}


@dataclass
class MonthMigrationResult:
    month: str
    source_sheet: str
    target_sheet: str
    days_expected: int
    source_day1_row: int | None = None
    source_total_row: int | None = None
    migrated_days: int = 0
    complete_days: int = 0
    partial_days: int = 0
    copied_cells: int = 0
    conflicts: list[str] = field(default_factory=list)
    conflict_counts: Counter = field(default_factory=Counter)
    notes: list[str] = field(default_factory=list)
    note_counts: Counter = field(default_factory=Counter)

    @property
    def status(self) -> str:
        if self.complete_days == self.days_expected and not self.conflicts:
            return "Listo"
        if self.migrated_days == 0:
            return "Pendiente"
        return "Validar"


def normalize_sheet_name(value: str) -> str:
    return "".join((value or "").upper().split())


def find_row_by_first_col(ws, expected) -> int | None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == expected:
            return row
    return None


def as_int(value) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def as_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def add_conflict(result: MonthMigrationResult, key: str, detail: str) -> None:
    result.conflicts.append(detail)
    result.conflict_counts[key] += 1


def add_note(result: MonthMigrationResult, key: str, detail: str) -> None:
    result.notes.append(detail)
    result.note_counts[key] += 1


def detect_summary_manual_values(source_formula_ws, result: MonthMigrationResult) -> None:
    labels = {
        "Ingresos",
        "Personas Atendidas",
        "Días Personas Atendidas",
        "Altas",
        "Indice ocupacional %",
        "Promedio día de estada",
        "fallecidos",
    }
    for row in range(1, source_formula_ws.max_row + 1):
        label = source_formula_ws.cell(row, 1).value
        value = source_formula_ws.cell(row, 3).value
        if label in labels and value not in (None, "") and not (isinstance(value, str) and value.startswith("=")):
            add_conflict(result, "manual_summary_value", f"Resumen manual sospechoso: {label} = {value}")


def month_key_map(sheet_names: Iterable[str]) -> dict[str, str]:
    return {normalize_sheet_name(name): name for name in sheet_names}


def build_target_sheet_map(sheet_names: Iterable[str]) -> dict[str, str]:
    return {
        normalize_sheet_name(name): name
        for name in sheet_names
        if any(name == spec.label for spec in MONTHS)
    }


def row_for_day(day: int) -> int:
    return 6 + day


def source_header_month_year(ws) -> tuple[str | None, int | None]:
    month_value = ws["P2"].value if ws.max_row >= 2 else None
    year_value = as_int(ws["R2"].value if ws.max_row >= 2 else None)
    month_name = str(month_value).strip().upper() if month_value not in (None, "") else None
    return month_name, year_value


def compare_header_to_sheet(source_formula_ws, spec: MonthSpec, result: MonthMigrationResult) -> None:
    month_name, year_value = source_header_month_year(source_formula_ws)
    if month_name or year_value:
        expected = spec.label
        if month_name not in expected or str(year_value) not in expected:
            add_conflict(
                result,
                "header_mismatch",
                f"Encabezado interno no coincide con la pestaña: {month_name} {year_value}",
            )


def migrate_month(
    spec: MonthSpec,
    source_formula_wb,
    source_value_wb,
    target_wb,
    source_sheet_map: dict[str, str],
    target_sheet_map: dict[str, str],
    default_camas_base: int,
) -> MonthMigrationResult:
    normalized = normalize_sheet_name(spec.label)
    source_sheet_name = source_sheet_map[normalized]
    target_sheet_name = target_sheet_map[normalized]

    source_formula_ws = source_formula_wb[source_sheet_name]
    source_value_ws = source_value_wb[source_sheet_name]
    target_ws = target_wb[target_sheet_name]

    result = MonthMigrationResult(
        month=spec.label,
        source_sheet=source_sheet_name,
        target_sheet=target_sheet_name,
        days_expected=spec.days,
    )

    compare_header_to_sheet(source_formula_ws, spec, result)
    detect_summary_manual_values(source_formula_ws, result)

    if spec.label in KNOWN_MANUAL_SUMMARY_MONTHS:
        add_conflict(result, "known_manual_summary_month", "Mes identificado previamente con resumen manual sospechoso")
    if spec.label == "MAYO 2025":
        add_conflict(result, "known_daily_balance_issue", "Mes identificado previamente con al menos un problema de balance diario")
    if spec.label in {"AGOSTO 2025", "SEPTIEMBRE 2025", "OCTUBRE 2025", "NOVIEMBRE 2025", "DICIEMBRE 2025", "ENERO 2026", "FEBRERO 2026"}:
        add_conflict(result, "legacy_structure_variant", "La estructura antigua cambia desde este mes; migrado por detección y no por posición fija")

    day1_row = find_row_by_first_col(source_value_ws, 1)
    total_row = find_row_by_first_col(source_value_ws, "Total")
    result.source_day1_row = day1_row
    result.source_total_row = total_row

    if day1_row is None or total_row is None or total_row <= day1_row:
        add_conflict(result, "structure_error", "No se pudo detectar correctamente el inicio de días o la fila Total")
        return result

    if total_row - day1_row != spec.days:
        add_conflict(
            result,
            "unexpected_day_count",
            f"Cantidad de filas diarias inesperada: detectadas {total_row - day1_row}, esperadas {spec.days}",
        )

    for source_row in range(day1_row, total_row):
        day = as_int(source_value_ws.cell(source_row, 1).value)
        if day is None or day < 1 or day > spec.days:
            add_conflict(result, "invalid_day_number", f"Fila {source_row} con día inválido: {source_value_ws.cell(source_row, 1).value}")
            continue

        target_row = row_for_day(day)
        day_conflicts_before = len(result.conflicts)

        if day == 1:
            source_initial_census = as_number(source_value_ws.cell(source_row, 2).value)
            if source_initial_census is None:
                add_conflict(result, "missing_day1_census", "Día 1 sin existencia inicial")
            else:
                target_ws[f"C{target_row}"] = source_initial_census
                result.copied_cells += 1

        income_components = {
            "urgencia": as_number(source_value_ws.cell(source_row, 3).value),
            "hospitalizacion": as_number(source_value_ws.cell(source_row, 4).value),
            "ambulatorio": as_number(source_value_ws.cell(source_row, 5).value),
            "ley_urgencia": as_number(source_value_ws.cell(source_row, 6).value),
            "ugcc": as_number(source_value_ws.cell(source_row, 7).value),
            "traslado_ingreso": as_number(source_value_ws.cell(source_row, 9).value),
        }
        source_procedencia_total = as_number(source_value_ws.cell(source_row, 8).value)
        source_total_ingresos = as_number(source_value_ws.cell(source_row, 10).value)

        if any(value is not None for value in income_components.values()):
            for key, target_col in TARGET_INPUT_COLUMN_MAP.items():
                if key not in income_components:
                    continue
                value = income_components[key]
                if value is not None:
                    target_ws[f"{target_col}{target_row}"] = value
                    result.copied_cells += 1

            component_sum = sum(value or 0 for value in income_components.values() if value is not None)
            procedencia_sum = sum(
                value or 0
                for key, value in income_components.items()
                if key != "traslado_ingreso" and value is not None
            )

            if source_procedencia_total is not None and procedencia_sum != source_procedencia_total:
                add_conflict(
                    result,
                    "income_procedencia_mismatch",
                    f"Día {day}: total según procedencia no calza ({procedencia_sum} vs {source_procedencia_total})",
                )
            if source_total_ingresos is not None and component_sum != source_total_ingresos:
                add_conflict(
                    result,
                    "income_total_mismatch",
                    f"Día {day}: total ingresos no calza con desglose ({component_sum} vs {source_total_ingresos})",
                )
        elif source_total_ingresos not in (None, 0):
            target_ws[f"{TARGET_INPUT_COLUMN_MAP['total_ingresos_informado']}{target_row}"] = source_total_ingresos
            result.copied_cells += 1
            add_note(
                result,
                "income_total_only",
                f"Día {day}: migrado usando solo total informado de ingresos ({source_total_ingresos})",
            )

        egress_components = {
            "alta_hogar": as_number(source_value_ws.cell(source_row, 11).value),
            "traslado_servicio": as_number(source_value_ws.cell(source_row, 12).value),
            "fallecidos": as_number(source_value_ws.cell(source_row, 13).value),
        }
        source_total_egresos = as_number(source_value_ws.cell(source_row, 14).value)
        if any(value is not None for value in egress_components.values()):
            for key, target_col in TARGET_INPUT_COLUMN_MAP.items():
                if key not in egress_components:
                    continue
                value = egress_components[key]
                if value is not None:
                    target_ws[f"{target_col}{target_row}"] = value
                    result.copied_cells += 1

            component_sum = sum(value or 0 for value in egress_components.values() if value is not None)
            if source_total_egresos is not None and component_sum != source_total_egresos:
                add_conflict(
                    result,
                    "egress_total_mismatch",
                    f"Día {day}: total egresos no calza con desglose ({component_sum} vs {source_total_egresos})",
                )
        elif source_total_egresos not in (None, 0):
            target_ws[f"{TARGET_INPUT_COLUMN_MAP['total_egresos_informado']}{target_row}"] = source_total_egresos
            result.copied_cells += 1
            add_note(
                result,
                "egress_total_only",
                f"Día {day}: migrado usando solo total informado de egresos ({source_total_egresos})",
            )

        for source_key in ("mismo_dia", "dias_estada_total", "dias_estada_benef"):
            source_value = as_number(source_value_ws.cell(source_row, SOURCE_STATUS_COLUMN_MAP[source_key]).value)
            if source_value is not None:
                target_ws[f"{TARGET_INPUT_COLUMN_MAP[source_key]}{target_row}"] = source_value
                result.copied_cells += 1

        source_camas_disponibles = as_number(source_value_ws.cell(source_row, SOURCE_STATUS_COLUMN_MAP["camas_disponibles"]).value)
        target_ws[f"R{target_row}"] = source_camas_disponibles if source_camas_disponibles is not None else default_camas_base
        result.copied_cells += 1

        source_prev = as_number(source_value_ws.cell(source_row, 2).value)
        source_q = as_number(source_value_ws.cell(source_row, 17).value)
        if day > 1 and source_prev is not None:
            prev_day_source_q = as_number(source_value_ws.cell(source_row - 1, 17).value)
            if prev_day_source_q is not None and source_prev != prev_day_source_q:
                add_conflict(
                    result,
                    "carry_forward_mismatch",
                    f"Día {day}: existencia inicial {source_prev} no coincide con camas ocupadas del día previo {prev_day_source_q}",
                )
        if source_prev is not None and source_total_ingresos is not None and source_total_egresos is not None and source_q is not None:
            expected_q = source_prev + source_total_ingresos - source_total_egresos
            if expected_q != source_q:
                add_conflict(
                    result,
                    "daily_balance_issue",
                    f"Día {day}: balance diario no calza ({source_prev} + {source_total_ingresos} - {source_total_egresos} = {expected_q}, origen muestra {source_q})",
                )

        result.migrated_days += 1
        if len(result.conflicts) == day_conflicts_before:
            result.complete_days += 1
        else:
            result.partial_days += 1

    return result


def update_migration_tracker(target_wb, result: MonthMigrationResult) -> None:
    ws = target_wb["03_MIGRACION"]
    row = 4 + MONTHS.index(next(spec for spec in MONTHS if spec.label == result.month))
    ws[f"B{row}"] = result.status
    ws[f"C{row}"] = f"{SOURCE_PATH.name} :: {result.source_sheet}"
    ws[f"D{row}"] = "Migracion automatica Codex"
    ws[f"E{row}"] = "Pendiente"
    ws[f"F{row}"] = (
        f"dias={result.migrated_days}/{result.days_expected}; "
        f"completos={result.complete_days}; parciales={result.partial_days}; "
        f"conflictos={len(result.conflicts)}; notas={len(result.notes)}"
    )


def write_csv_report(results: list[MonthMigrationResult]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "month",
                "source_sheet",
                "target_sheet",
                "days_expected",
                "source_day1_row",
                "source_total_row",
                "migrated_days",
                "complete_days",
                "partial_days",
                "copied_cells",
                "conflicts_count",
                "notes_count",
                "status",
                "conflict_summary",
                "note_summary",
            ]
        )
        for result in results:
            summary = "; ".join(
                f"{key}={count}"
                for key, count in sorted(result.conflict_counts.items())
            )
            note_summary = "; ".join(
                f"{key}={count}"
                for key, count in sorted(result.note_counts.items())
            )
            writer.writerow(
                [
                    result.month,
                    result.source_sheet,
                    result.target_sheet,
                    result.days_expected,
                    result.source_day1_row,
                    result.source_total_row,
                    result.migrated_days,
                    result.complete_days,
                    result.partial_days,
                    result.copied_cells,
                    len(result.conflicts),
                    len(result.notes),
                    result.status,
                    summary,
                    note_summary,
                ]
            )


def write_markdown_report(results: list[MonthMigrationResult]) -> None:
    total_conflicts = sum(len(result.conflicts) for result in results)
    total_notes = sum(len(result.notes) for result in results)
    total_complete = sum(result.complete_days for result in results)
    total_partial = sum(result.partial_days for result in results)
    total_days = sum(result.days_expected for result in results)

    lines = [
        "# Reporte de migración HODOM -> estructura ideal",
        "",
        f"- Fuente: `{SOURCE_PATH}`",
        f"- Plantilla base: `{TEMPLATE_PATH}`",
        f"- Workbook migrado: `{OUTPUT_WORKBOOK}`",
        f"- Días esperados: `{total_days}`",
        f"- Días completos: `{total_complete}`",
        f"- Días parciales: `{total_partial}`",
        f"- Conflictos totales: `{total_conflicts}`",
        f"- Notas informativas: `{total_notes}`",
        "",
        "## Resumen por mes",
        "",
    ]

    for result in results:
        lines.append(f"### {result.month}")
        lines.append(f"- Estado: `{result.status}`")
        lines.append(
            f"- Migración: `{result.migrated_days}/{result.days_expected}` días, "
            f"`{result.complete_days}` completos, `{result.partial_days}` parciales, `{result.copied_cells}` celdas copiadas"
        )
        if result.conflict_counts:
            summary = ", ".join(
                f"`{key}` x {count}"
                for key, count in sorted(result.conflict_counts.items())
            )
            lines.append(f"- Conflictos: {summary}")
        else:
            lines.append("- Conflictos: ninguno")

        if result.note_counts:
            note_summary = ", ".join(
                f"`{key}` x {count}"
                for key, count in sorted(result.note_counts.items())
            )
            lines.append(f"- Notas: {note_summary}")

        if result.conflicts:
            lines.append("- Detalle:")
            for detail in result.conflicts[:10]:
                lines.append(f"  - {detail}")
            if len(result.conflicts) > 10:
                lines.append(f"  - ... y {len(result.conflicts) - 10} más")
        elif result.notes:
            lines.append("- Detalle:")
            for detail in result.notes[:10]:
                lines.append(f"  - {detail}")
            if len(result.notes) > 10:
                lines.append(f"  - ... y {len(result.notes) - 10} más")
        lines.append("")

    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    source_formula_wb = load_workbook(SOURCE_PATH, data_only=False)
    source_value_wb = load_workbook(SOURCE_PATH, data_only=True)
    target_wb = load_workbook(TEMPLATE_PATH, data_only=False)

    source_sheet_map = month_key_map(source_value_wb.sheetnames)
    target_sheet_map = build_target_sheet_map(target_wb.sheetnames)
    default_camas_base = as_int(target_wb["01_PARAMETROS"]["B5"].value) or 16

    results: list[MonthMigrationResult] = []
    for spec in MONTHS:
        normalized = normalize_sheet_name(spec.label)
        if normalized not in source_sheet_map:
            result = MonthMigrationResult(
                month=spec.label,
                source_sheet="",
                target_sheet=spec.label,
                days_expected=spec.days,
            )
            add_conflict(result, "missing_source_sheet", "No se encontró la hoja de origen correspondiente")
            results.append(result)
            continue
        result = migrate_month(
            spec=spec,
            source_formula_wb=source_formula_wb,
            source_value_wb=source_value_wb,
            target_wb=target_wb,
            source_sheet_map=source_sheet_map,
            target_sheet_map=target_sheet_map,
            default_camas_base=default_camas_base,
        )
        update_migration_tracker(target_wb, result)
        results.append(result)

    if hasattr(target_wb, "calculation"):
        target_wb.calculation.fullCalcOnLoad = True
        target_wb.calculation.forceFullCalc = True

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    target_wb.save(OUTPUT_WORKBOOK)
    write_csv_report(results)
    write_markdown_report(results)

    print(OUTPUT_WORKBOOK)
    print(OUTPUT_CSV)
    print(OUTPUT_MD)


if __name__ == "__main__":
    main()
