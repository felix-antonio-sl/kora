#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time as dt_time, timedelta
from pathlib import Path
from statistics import mean
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
TMP_DIR = ROOT / "tmp" / "spreadsheets"
GEOCODE_CACHE_PATH = TMP_DIR / "hd_routes_geocode_cache.json"
WORKBOOK_PATH = OUTPUT_DIR / "hd_routes_operational_analysis_2026-01-01_to_2026-03-24.xlsx"
REPORT_PATH = OUTPUT_DIR / "hd_routes_operational_analysis_2026-01-01_to_2026-03-24.md"

GPS_SOURCE = Path(
    "/Users/felixsanhueza/Downloads/drives_stops_report_2026_01_01_00_00_00_2026_03_25_00_00_00_1774372482.csv"
)
PLANNING_SOURCES = [
    Path("/Users/felixsanhueza/Downloads/ENERO 2026.xlsx"),
    Path("/Users/felixsanhueza/Downloads/FEBRERO 2026.xlsx"),
    Path("/Users/felixsanhueza/Downloads/MARZO 2026.xlsx"),
]

COMPARABLE_START = date(2026, 1, 1)
COMPARABLE_END = date(2026, 3, 24)
ANNEX_ONLY_DATE = date(2026, 3, 25)

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
ARCGIS_URL = "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates"
USER_AGENT = "kora-hd-route-analysis/1.0 (local execution)"
REQUEST_TIMEOUT = 15
REQUEST_DELAY_SECONDS = 1.05

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

LOCALITY_HINTS = [
    "SAN CARLOS",
    "NIQUEN",
    "SAN NICOLAS",
    "SAN GREGORIO",
    "CACHAPOAL",
    "COIHUECO",
]

ADDRESS_REPLACEMENTS = {
    r"\bPSJE\b": "PASAJE",
    r"\bPJE\b": "PASAJE",
    r"\bPSJ\b": "PASAJE",
    r"\bPOB\.\b": "POBLACION",
    r"\bPOBL\.\b": "POBLACION",
    r"\bAV\.\b": "AVENIDA",
    r"\bVDA\.\b": "VEREDA",
    r"\bSTA\b": "SANTA",
    r"\bS/N\b": "SN",
}


@dataclass
class GeocodeResult:
    direccion_normalizada: str
    lat: float | None
    lon: float | None
    fuente_geocoder: str
    calidad_geocoder: str
    cache_key: str
    consulta_usada: str
    display_name: str
    token_overlap: float

    def as_row(self) -> dict[str, Any]:
        return {
            "direccion_normalizada": self.direccion_normalizada,
            "lat": self.lat,
            "lon": self.lon,
            "fuente_geocoder": self.fuente_geocoder,
            "calidad_geocoder": self.calidad_geocoder,
            "cache_key": self.cache_key,
            "consulta_usada": self.consulta_usada,
            "display_name": self.display_name,
            "token_overlap": round(self.token_overlap, 3),
        }


def ensure_dirs() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, dt_time):
        return value.strftime("%H:%M")
    return str(value).strip()


def ascii_fold(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip(" ,")


def normalize_name(value: str) -> str:
    return normalize_spaces(ascii_fold(value).upper())


def normalize_address(value: str) -> str:
    text = normalize_spaces(ascii_fold(value).upper())
    text = text.replace('"', " ")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("'", " ")
    for pattern, replacement in ADDRESS_REPLACEMENTS.items():
        text = re.sub(pattern, replacement, text)
    text = re.sub(r"\s*-\s*", " ", text)
    text = normalize_spaces(text)
    return text


def parse_sheet_date(sheet_name: str, first_cell: Any) -> date:
    if isinstance(first_cell, datetime):
        return first_cell.date()
    if isinstance(first_cell, date):
        return first_cell
    match = re.search(r"(\d{2})\.(\d{2})", sheet_name.strip())
    if not match:
        raise ValueError(f"No pude inferir fecha de la hoja {sheet_name!r}")
    day_value = int(match.group(1))
    month_value = int(match.group(2))
    return date(2026, month_value, day_value)


def parse_scheduled_time(value: Any) -> dt_time | None:
    if isinstance(value, dt_time):
        return value
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"(\d{1,2}):(\d{2})", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return dt_time(hour=hour, minute=minute)
    return None


def is_header_row(values: list[Any]) -> bool:
    return clean_text(values[1]).upper() == "HORA"


def row_is_blank(values: list[Any]) -> bool:
    return all(clean_text(v) == "" for v in values[:11])


def leader_from_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, int, float)):
        return None
    text = normalize_spaces(str(value))
    if not text or text in {"-", "0", "0.0"}:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}( 00:00:00)?", text):
        return None
    return text


def parse_duration(text: str) -> int:
    if not text:
        return 0
    total = 0
    hours_match = re.search(r"(\d+)h", text)
    minutes_match = re.search(r"(\d+)min", text)
    seconds_match = re.search(r"(\d+)s", text)
    if hours_match:
        total += int(hours_match.group(1)) * 3600
    if minutes_match:
        total += int(minutes_match.group(1)) * 60
    if seconds_match:
        total += int(seconds_match.group(1))
    return total


def parse_distance_km(text: str) -> float | None:
    match = re.search(r"([\d.,]+)\s*Km", text or "")
    if not match:
        return None
    return float(match.group(1).replace(",", "."))


def parse_speed_kph(text: str) -> float | None:
    match = re.search(r"([\d.,]+)\s*kph", text or "", re.IGNORECASE)
    if not match:
        return None
    return float(match.group(1).replace(",", "."))


def parse_lat_lon(text: str) -> tuple[float | None, float | None]:
    match = re.search(r"(-?[\d.]+)\s*,\s*(-?[\d.]+)", text or "")
    if not match:
        return None, None
    return float(match.group(1)), float(match.group(2))


def dt_to_iso(value: datetime | None) -> str:
    return value.isoformat(sep=" ") if value else ""


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = (
        math.sin(delta_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius * c


def address_tokens(value: str) -> set[str]:
    folded = ascii_fold(value).upper()
    tokens = set(re.findall(r"[A-Z0-9]+", folded))
    stopwords = {
        "SAN",
        "CARLOS",
        "CHILE",
        "NUBLE",
        "REGION",
        "DE",
        "DEL",
        "LA",
        "EL",
        "LOS",
        "LAS",
        "SECTOR",
        "POBLACION",
        "VILLA",
        "KM",
        "SN",
        "RUTA",
        "CAMINO",
    }
    return {token for token in tokens if token not in stopwords and len(token) > 1}


def token_overlap(left: str, right: str) -> float:
    left_tokens = address_tokens(left)
    right_tokens = address_tokens(right)
    if not left_tokens:
        return 0.0
    return len(left_tokens & right_tokens) / len(left_tokens)


def detect_locality(normalized_address: str) -> str:
    for locality in LOCALITY_HINTS:
        if locality in normalized_address:
            return locality
    return "SAN CARLOS"


def build_query_candidates(normalized_address: str) -> list[str]:
    queries: list[str] = []

    def add(query: str) -> None:
        query = normalize_spaces(query)
        if query and query not in queries:
            queries.append(query)

    locality = detect_locality(normalized_address)
    region = "Region de Nuble, Chile"
    parts = [part.strip() for part in normalized_address.split(",") if part.strip()]
    base = ", ".join(parts) if parts else normalized_address
    add(f"{base}, {region}")

    without_sn = normalize_spaces(re.sub(r"\bSN\b", " ", base))
    if without_sn != base:
        add(f"{without_sn}, {region}")

    if len(parts) >= 2:
        first = parts[0]
        second = parts[1]
        numbers = " ".join(re.findall(r"\b\d+[A-Z]?\b", second))
        if numbers:
            add(f"{first} {numbers}, {locality}, {region}")
        second_words = normalize_spaces(re.sub(r"\b(PASAJE|AVENIDA|VEREDA)\b", " ", second))
        if second_words:
            add(f"{first} {second_words}, {locality}, {region}")
        add(f"{first}, {locality}, {region}")
    elif parts:
        add(f"{parts[0]}, {locality}, {region}")

    if "ELEAM SAN AGUSTIN" in normalized_address:
        add(f"El Cape, {locality}, {region}")
        add(f"San Agustin, El Cape, {locality}, {region}")

    if "LAS ARBOLEDAS" in normalized_address:
        add(f"Las Arboledas, {locality}, {region}")

    if "27 DE ABRIL" in normalized_address:
        numbers = " ".join(re.findall(r"\b\d+[A-Z]?\b", normalized_address))
        if numbers:
            add(f"27 de Abril {numbers}, {locality}, {region}")
        add(f"Roberto Bustamante, Villa 27 de Abril, {locality}, {region}")

    if "SAN CAMILO" in normalized_address:
        add(f"San Camilo, {locality}, {region}")

    if "AGUA BUENA" in normalized_address:
        add(f"Agua Buena, {locality}, {region}")

    if "CUAGRAPANGUE" in normalized_address:
        add(f"{normalized_address.replace('CUAGRAPANGUE', 'CUADRAPANGUE')}, {region}")

    if "L DE CHILE" in normalized_address:
        add(f"{normalized_address.replace('L DE CHILE', 'LAGO DE CHILE')}, {region}")

    return queries


def load_geocode_cache() -> dict[str, dict[str, Any]]:
    if not GEOCODE_CACHE_PATH.exists():
        return {}
    return json.loads(GEOCODE_CACHE_PATH.read_text(encoding="utf-8"))


def save_geocode_cache(cache: dict[str, dict[str, Any]]) -> None:
    GEOCODE_CACHE_PATH.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def select_geocode_result(
    normalized_address: str, query: str, results: list[dict[str, Any]]
) -> dict[str, Any] | None:
    scored: list[tuple[float, dict[str, Any]]] = []
    for result in results:
        display = result.get("display_name", "")
        overlap = token_overlap(normalized_address, display)
        importance = float(result.get("importance", 0.0) or 0.0)
        score = overlap * 5 + importance
        scored.append((score, result))
    if not scored:
        return None
    scored.sort(key=lambda item: item[0], reverse=True)
    _, best = scored[0]
    best_overlap = token_overlap(normalized_address, best.get("display_name", ""))
    if best_overlap < 0.12 and len(results) == 1:
        # Keep low-confidence single hits as fallback for sparse rural addresses.
        return best
    if best_overlap < 0.12:
        return None
    return best


def select_arcgis_candidate(
    normalized_address: str, candidates: list[dict[str, Any]]
) -> dict[str, Any] | None:
    scored: list[tuple[float, dict[str, Any]]] = []
    for candidate in candidates:
        address = candidate.get("address", "")
        overlap = token_overlap(normalized_address, address)
        score = float(candidate.get("score", 0.0) or 0.0)
        ranked = overlap * 5 + (score / 100)
        scored.append((ranked, candidate))
    if not scored:
        return None
    scored.sort(key=lambda item: item[0], reverse=True)
    _, best = scored[0]
    overlap = token_overlap(normalized_address, best.get("address", ""))
    score = float(best.get("score", 0.0) or 0.0)
    if overlap < 0.08 and score < 84:
        return None
    return best


def geocode_addresses(addresses: list[str]) -> dict[str, GeocodeResult]:
    cache = load_geocode_cache()
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    resolved: dict[str, GeocodeResult] = {}
    requests_made = 0

    for index, normalized_address in enumerate(addresses, start=1):
        cache_key = hashlib.sha1(normalized_address.encode("utf-8")).hexdigest()
        cached = cache.get(cache_key)
        if cached and cached.get("lat") is not None:
            resolved[normalized_address] = GeocodeResult(**cached)
            continue

        best_result: GeocodeResult | None = None
        for query in build_query_candidates(normalized_address):
            try:
                response = session.get(
                    NOMINATIM_URL,
                    params={
                        "q": query,
                        "format": "jsonv2",
                        "limit": 3,
                        "countrycodes": "cl",
                        "addressdetails": 1,
                    },
                    timeout=REQUEST_TIMEOUT,
                )
                requests_made += 1
                payload = response.json()
            except requests.RequestException:
                time.sleep(REQUEST_DELAY_SECONDS)
                continue

            selected = select_geocode_result(normalized_address, query, payload)
            time.sleep(REQUEST_DELAY_SECONDS)
            if not selected:
                continue
            display_name = selected.get("display_name", "")
            overlap = token_overlap(normalized_address, display_name)
            quality = "alta" if overlap >= 0.55 else "media" if overlap >= 0.25 else "baja"
            best_result = GeocodeResult(
                direccion_normalizada=normalized_address,
                lat=float(selected["lat"]),
                lon=float(selected["lon"]),
                fuente_geocoder="nominatim",
                calidad_geocoder=quality,
                cache_key=cache_key,
                consulta_usada=query,
                display_name=display_name,
                token_overlap=overlap,
            )
            break

        if best_result is None:
            for query in build_query_candidates(normalized_address)[:3]:
                try:
                    response = session.get(
                        ARCGIS_URL,
                        params={
                            "SingleLine": query,
                            "f": "pjson",
                            "maxLocations": 3,
                            "countryCode": "CHL",
                            "forStorage": "false",
                            "outFields": "Match_addr,Addr_type",
                        },
                        timeout=REQUEST_TIMEOUT,
                    )
                    requests_made += 1
                    payload = response.json()
                except requests.RequestException:
                    continue

                candidates = payload.get("candidates", [])
                selected = select_arcgis_candidate(normalized_address, candidates)
                if not selected:
                    continue
                location = selected.get("location") or {}
                display_name = selected.get("address", "")
                overlap = token_overlap(normalized_address, display_name)
                quality = "alta" if overlap >= 0.55 else "media" if overlap >= 0.2 else "baja"
                best_result = GeocodeResult(
                    direccion_normalizada=normalized_address,
                    lat=location.get("y"),
                    lon=location.get("x"),
                    fuente_geocoder="arcgis",
                    calidad_geocoder=quality,
                    cache_key=cache_key,
                    consulta_usada=query,
                    display_name=display_name,
                    token_overlap=overlap,
                )
                break

        if best_result is None:
            best_result = GeocodeResult(
                direccion_normalizada=normalized_address,
                lat=None,
                lon=None,
                fuente_geocoder="nominatim",
                calidad_geocoder="sin_match",
                cache_key=cache_key,
                consulta_usada="",
                display_name="",
                token_overlap=0.0,
            )

        cache[cache_key] = best_result.__dict__
        resolved[normalized_address] = best_result
        save_geocode_cache(cache)
        if index % 10 == 0 or index == len(addresses):
            resolved_count = sum(1 for item in cache.values() if item.get("lat") is not None)
            print(
                f"Geocoding progreso: {index}/{len(addresses)} direcciones, "
                f"resueltas={resolved_count}, consultas={requests_made}"
            )

    save_geocode_cache(cache)
    print(
        f"Geocoding completado: {len(addresses)} direcciones, "
        f"{requests_made} consultas HTTP, cache={GEOCODE_CACHE_PATH}"
    )
    return resolved


def parse_planned_visits() -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, Any]]]:
    visits: list[dict[str, Any]] = []
    outliers: list[dict[str, Any]] = []
    block_meta: dict[str, dict[str, Any]] = {}

    for workbook_path in PLANNING_SOURCES:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_date = parse_sheet_date(sheet_name, worksheet["A1"].value)
            block_sequence = 0
            current_block_id: str | None = None
            current_leader: str | None = None
            comparable = COMPARABLE_START <= sheet_date <= COMPARABLE_END
            annex_only = sheet_date == ANNEX_ONLY_DATE

            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = list(row[:11]) + [None] * max(0, 11 - len(row))
                if row_is_blank(values):
                    current_block_id = None
                    current_leader = None
                    continue
                if is_header_row(values):
                    current_block_id = None
                    current_leader = None
                    continue

                scheduled_time = parse_scheduled_time(values[1])
                patient = normalize_spaces(clean_text(values[7]))
                visit_type = normalize_spaces(clean_text(values[8]))
                address = normalize_spaces(clean_text(values[9]))
                phone = normalize_spaces(clean_text(values[10]))

                leader = leader_from_cell(values[0])
                if not leader and current_block_id is None and scheduled_time and patient:
                    for index in range(2, 7):
                        leader = leader_from_cell(values[index])
                        if leader:
                            break
                if leader:
                    if current_leader != leader or current_block_id is None:
                        block_sequence += 1
                        current_leader = leader
                        current_block_id = f"{sheet_date.isoformat()}_B{block_sequence:02d}"
                        block_meta[current_block_id] = {
                            "fecha": sheet_date.isoformat(),
                            "block_id": current_block_id,
                            "block_sequence": block_sequence,
                            "lider_bloque": current_leader,
                            "sheet_name": sheet_name,
                            "source_file": workbook_path.name,
                            "visit_ids": [],
                            "scheduled_times": [],
                        }
                elif current_block_id is None:
                    continue

                if scheduled_time and patient and comparable:
                    scheduled_dt = datetime.combine(sheet_date, scheduled_time)
                    visit_id = f"{sheet_date.isoformat()}_{sheet_name}_{row_index:03d}"
                    normalized_address = normalize_address(address)
                    visit = {
                        "visita_programada_id": visit_id,
                        "fecha": sheet_date.isoformat(),
                        "hoja": sheet_name,
                        "archivo_fuente": workbook_path.name,
                        "fila_fuente": row_index,
                        "bloque_diario": current_block_id,
                        "block_sequence": block_meta[current_block_id]["block_sequence"],
                        "lider_bloque": current_leader or "",
                        "hora_programada": scheduled_time.strftime("%H:%M"),
                        "horario_programado": scheduled_dt,
                        "medico": normalize_spaces(clean_text(values[2])),
                        "fono": normalize_spaces(clean_text(values[3])),
                        "kine": normalize_spaces(clean_text(values[4])),
                        "enfermera": normalize_spaces(clean_text(values[5])),
                        "tens": normalize_spaces(clean_text(values[6])),
                        "paciente": patient,
                        "paciente_id": "",
                        "tipo_visita": visit_type,
                        "direccion": address,
                        "direccion_normalizada": normalized_address,
                        "telefono": phone,
                        "lat": None,
                        "lon": None,
                        "geocode_quality": "",
                    }
                    visits.append(visit)
                    block_meta[current_block_id]["visit_ids"].append(visit_id)
                    block_meta[current_block_id]["scheduled_times"].append(scheduled_dt)
                else:
                    has_planning_payload = any(
                        normalize_spaces(clean_text(values[index]))
                        for index in range(2, 11)
                    ) or patient
                    if has_planning_payload:
                        if comparable:
                            outliers.append(
                                {
                                    "category": "planned_row_incomplete",
                                    "fecha": sheet_date.isoformat(),
                                    "dispositivo": "",
                                    "bloque_diario": current_block_id or "",
                                    "visita_programada_id": "",
                                    "reference": f"{workbook_path.name}:{sheet_name}:{row_index}",
                                    "detail": (
                                        "Fila programada fuera de comparacion automatica "
                                        f"(hora={clean_text(values[1])!r}, paciente={patient!r})"
                                    ),
                                }
                            )
                        elif annex_only:
                            outliers.append(
                                {
                                    "category": "annex_2026-03-25",
                                    "fecha": sheet_date.isoformat(),
                                    "dispositivo": "",
                                    "bloque_diario": current_block_id or "",
                                    "visita_programada_id": "",
                                    "reference": f"{workbook_path.name}:{sheet_name}:{row_index}",
                                    "detail": "Programacion disponible sin telemetria comparable para 2026-03-25.",
                                }
                            )

    patient_index: dict[str, str] = {}
    for counter, patient_name in enumerate(
        sorted({normalize_name(item["paciente"]) for item in visits}), start=1
    ):
        patient_index[patient_name] = f"PAC-{counter:03d}"

    for visit in visits:
        visit["paciente_id"] = patient_index[normalize_name(visit["paciente"])]

    return visits, outliers, block_meta


def parse_gps_events() -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    current_device = ""
    event_counter = 0

    with GPS_SOURCE.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if not row:
                continue
            marker = clean_text(row[0])
            if marker == "Dispositivo:":
                current_device = clean_text(row[1])
                continue
            if marker not in {"Detenido", "Movimiento"} or not current_device:
                continue
            try:
                start = datetime.fromisoformat(clean_text(row[1]))
                end = datetime.fromisoformat(clean_text(row[2]))
            except ValueError:
                continue
            if start.date() < COMPARABLE_START or start.date() > COMPARABLE_END:
                continue
            event_counter += 1
            duration_seconds = parse_duration(clean_text(row[3]))
            lat, lon = parse_lat_lon(clean_text(row[6])) if marker == "Detenido" else (None, None)
            event = {
                "event_id": f"GPS-{event_counter:05d}",
                "stop_id": f"STOP-{event_counter:05d}" if marker == "Detenido" else "",
                "device": current_device,
                "date": start.date().isoformat(),
                "tipo_evento": marker,
                "inicio": start,
                "fin": end,
                "duracion_seg": duration_seconds,
                "ralenti_seg": parse_duration(clean_text(row[4])),
                "conductor": clean_text(row[5]),
                "lat": lat,
                "lon": lon,
                "distancia_km": parse_distance_km(clean_text(row[6])) if marker == "Movimiento" else None,
                "vel_max_kph": parse_speed_kph(clean_text(row[7])) if marker == "Movimiento" else None,
                "vel_media_kph": parse_speed_kph(clean_text(row[8])) if marker == "Movimiento" else None,
                "candidate_stop": marker == "Detenido" and duration_seconds >= 300,
                "base_stop": False,
            }
            events.append(event)

    return events


def build_base_clusters(events: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    by_device: dict[str, list[dict[str, Any]]] = defaultdict(list)
    stop_by_date: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        if event["tipo_evento"] == "Detenido" and event["lat"] is not None:
            stop_by_date[(event["device"], event["date"])].append(event)
    for (device, _), daily_stops in stop_by_date.items():
        ordered = sorted(daily_stops, key=lambda item: item["inicio"])
        if ordered:
            by_device[device].append(ordered[0])
            by_device[device].append(ordered[-1])

    clusters_by_device: dict[str, dict[str, Any]] = {}
    for device, points in by_device.items():
        clusters: list[dict[str, Any]] = []
        for point in points:
            assigned = False
            for cluster in clusters:
                distance = haversine_m(
                    point["lat"], point["lon"], cluster["lat"], cluster["lon"]
                )
                if distance <= 350:
                    cluster["points"].append(point)
                    cluster["lat"] = mean(item["lat"] for item in cluster["points"])
                    cluster["lon"] = mean(item["lon"] for item in cluster["points"])
                    assigned = True
                    break
            if not assigned:
                clusters.append({"lat": point["lat"], "lon": point["lon"], "points": [point]})
        base_cluster = max(clusters, key=lambda item: len(item["points"]))
        clusters_by_device[device] = {
            "lat": base_cluster["lat"],
            "lon": base_cluster["lon"],
            "samples": len(base_cluster["points"]),
        }

    for event in events:
        if event["tipo_evento"] != "Detenido" or event["lat"] is None:
            continue
        base = clusters_by_device.get(event["device"])
        if not base:
            continue
        event["base_stop"] = (
            haversine_m(event["lat"], event["lon"], base["lat"], base["lon"]) <= 350
        )

    return clusters_by_device


def enrich_visits_with_geocodes(
    visits: list[dict[str, Any]], geocodes: dict[str, GeocodeResult], outliers: list[dict[str, Any]]
) -> None:
    for visit in visits:
        result = geocodes.get(visit["direccion_normalizada"])
        if not result:
            continue
        visit["lat"] = result.lat
        visit["lon"] = result.lon
        visit["geocode_quality"] = result.calidad_geocoder
        if result.lat is None or result.lon is None:
            outliers.append(
                {
                    "category": "geocode_failed",
                    "fecha": visit["fecha"],
                    "dispositivo": "",
                    "bloque_diario": visit["bloque_diario"],
                    "visita_programada_id": visit["visita_programada_id"],
                    "reference": visit["direccion_normalizada"],
                    "detail": "Direccion sin resolucion geografica; no se fuerza matching espacial.",
                }
            )


def group_visits_by_block(visits: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for visit in visits:
        grouped[visit["bloque_diario"]].append(visit)
    for block_id in grouped:
        grouped[block_id].sort(key=lambda item: item["horario_programado"])
    return grouped


def block_windows(block_visits: dict[str, list[dict[str, Any]]]) -> dict[str, tuple[datetime, datetime]]:
    windows: dict[str, tuple[datetime, datetime]] = {}
    for block_id, visits in block_visits.items():
        starts = [visit["horario_programado"] for visit in visits]
        windows[block_id] = (min(starts), max(starts))
    return windows


def build_daily_index(events: list[dict[str, Any]]) -> tuple[
    dict[str, dict[str, list[dict[str, Any]]]],
    dict[str, dict[str, list[dict[str, Any]]]],
]:
    candidate_stops: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    events_by_day: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for event in events:
        events_by_day[event["date"]][event["device"]].append(event)
        if event["candidate_stop"] and not event["base_stop"]:
            candidate_stops[event["date"]][event["device"]].append(event)
    for date_key in events_by_day:
        for device in events_by_day[date_key]:
            events_by_day[date_key][device].sort(key=lambda item: item["inicio"])
    for date_key in candidate_stops:
        for device in candidate_stops[date_key]:
            candidate_stops[date_key][device].sort(key=lambda item: item["inicio"])
    return candidate_stops, events_by_day


def compute_pair_metrics(
    block_id: str,
    visits: list[dict[str, Any]],
    stops: list[dict[str, Any]],
    device_events: list[dict[str, Any]],
) -> dict[str, Any]:
    geocoded_visits = [visit for visit in visits if visit["lat"] is not None and visit["lon"] is not None]
    active_events = [event for event in device_events if not event["base_stop"]]
    if active_events:
        first_active = min(event["inicio"] for event in active_events)
        last_active = max(event["fin"] for event in active_events)
    else:
        first_active = None
        last_active = None

    high = 0
    medium = 0
    candidate_count = 0
    impossible = 0
    min_candidate_score: float | None = None

    for visit in geocoded_visits:
        visit_time = visit["horario_programado"]
        if first_active and visit_time < first_active - timedelta(minutes=30):
            impossible += 1
        elif last_active and visit_time > last_active + timedelta(minutes=30):
            impossible += 1

        best_distance = None
        best_delta = None
        for stop in stops:
            distance_m = haversine_m(visit["lat"], visit["lon"], stop["lat"], stop["lon"])
            delta_min = abs((stop["inicio"] - visit_time).total_seconds()) / 60
            if distance_m <= 800 and delta_min <= 90:
                best_distance = distance_m
                best_delta = delta_min
                break
            if distance_m <= 1500 and delta_min <= 150:
                if best_distance is None or distance_m + delta_min < best_distance + (best_delta or 0):
                    best_distance = distance_m
                    best_delta = delta_min
        if best_distance is None or best_delta is None:
            continue
        candidate_count += 1
        candidate_cost = best_distance / 500 + best_delta / 30
        min_candidate_score = candidate_cost if min_candidate_score is None else min(
            min_candidate_score, candidate_cost
        )
        if best_distance <= 800 and best_delta <= 90:
            high += 1
        elif best_distance <= 1500 and best_delta <= 150:
            medium += 1

    score = high * 3 + medium - impossible * 2
    return {
        "block_id": block_id,
        "score": score,
        "high_candidates": high,
        "medium_candidates": medium,
        "candidate_count": candidate_count,
        "impossible_overlaps": impossible,
        "geocoded_visits": len(geocoded_visits),
        "best_candidate_cost": round(min_candidate_score, 3) if min_candidate_score is not None else None,
    }


def blocks_overlap(
    left: tuple[datetime, datetime], right: tuple[datetime, datetime], buffer_minutes: int = 15
) -> bool:
    left_start, left_end = left
    right_start, right_end = right
    buffer_delta = timedelta(minutes=buffer_minutes)
    return not (left_end + buffer_delta <= right_start or right_end + buffer_delta <= left_start)


def assign_blocks_to_devices(
    date_key: str,
    block_visits: dict[str, list[dict[str, Any]]],
    block_meta: dict[str, dict[str, Any]],
    candidate_stops_by_day: dict[str, dict[str, list[dict[str, Any]]]],
    events_by_day: dict[str, dict[str, list[dict[str, Any]]]],
) -> list[dict[str, Any]]:
    blocks = [block_id for block_id in block_visits if block_id.startswith(date_key)]
    blocks.sort(key=lambda item: block_visits[item][0]["horario_programado"])
    devices = sorted(events_by_day.get(date_key, {}).keys())
    pair_scores: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    windows = block_windows(block_visits)

    for block_id in blocks:
        for device in devices:
            pair_scores[block_id][device] = compute_pair_metrics(
                block_id,
                block_visits[block_id],
                candidate_stops_by_day.get(date_key, {}).get(device, []),
                events_by_day.get(date_key, {}).get(device, []),
            )

    best_total = float("-inf")
    best_assignment: dict[str, str | None] = {}

    def search(
        index: int,
        current_assignment: dict[str, str | None],
        device_load: dict[str, list[str]],
        current_score: float,
    ) -> None:
        nonlocal best_total, best_assignment
        if index >= len(blocks):
            if current_score > best_total:
                best_total = current_score
                best_assignment = current_assignment.copy()
            return

        block_id = blocks[index]
        search(index + 1, {**current_assignment, block_id: None}, device_load, current_score)

        for device in devices:
            overlap = any(
                blocks_overlap(windows[block_id], windows[other_block])
                for other_block in device_load[device]
            )
            if overlap:
                continue
            pair_score = pair_scores[block_id][device]["score"]
            next_device_load = {key: value[:] for key, value in device_load.items()}
            next_device_load[device].append(block_id)
            search(
                index + 1,
                {**current_assignment, block_id: device},
                next_device_load,
                current_score + pair_score,
            )

    search(0, {}, defaultdict(list), 0.0)

    assignments: list[dict[str, Any]] = []
    for block_id in blocks:
        selected_device = best_assignment.get(block_id)
        selected_score = pair_scores[block_id].get(selected_device, {}).get("score", 0) if selected_device else 0
        alternative_scores = sorted(
            [
                score["score"]
                for device, score in pair_scores[block_id].items()
                if device != selected_device
            ],
            reverse=True,
        )
        second_best = alternative_scores[0] if alternative_scores else 0
        candidate_count = (
            pair_scores[block_id][selected_device]["candidate_count"] if selected_device else 0
        )
        if selected_device is None or selected_score <= 0:
            confidence = "ambiguo"
            reason = "Sin evidencia operacional suficiente para asignar dispositivo."
        else:
            if second_best <= 0:
                margin_ratio = 1.0
            else:
                margin_ratio = (selected_score - second_best) / abs(second_best)
            if candidate_count < 2 or margin_ratio < 0.2:
                confidence = "ambiguo"
                reason = (
                    f"Margen insuficiente frente a la segunda mejor opcion (score={selected_score}, second={second_best})."
                )
            elif margin_ratio >= 0.5 and selected_score >= 4:
                confidence = "asignado con confianza"
                reason = ""
            else:
                confidence = "asignado con reserva"
                reason = (
                    f"Asignacion util pero con holgura limitada (score={selected_score}, second={second_best})."
                )
        metrics = pair_scores[block_id].get(selected_device, {}) if selected_device else {}
        assignments.append(
            {
                "fecha": date_key,
                "bloque_diario": block_id,
                "block_sequence": block_meta[block_id]["block_sequence"],
                "lider_bloque": block_meta[block_id]["lider_bloque"],
                "dispositivo_asignado": selected_device or "",
                "score_asignacion": selected_score,
                "confianza_asignacion": confidence,
                "motivo_ambiguedad": reason,
                "second_best_score": second_best,
                "candidate_count": candidate_count,
                "high_candidates": metrics.get("high_candidates", 0),
                "medium_candidates": metrics.get("medium_candidates", 0),
                "impossible_overlaps": metrics.get("impossible_overlaps", 0),
            }
        )
    return assignments


def match_assigned_visits(
    visits: list[dict[str, Any]],
    assignments: list[dict[str, Any]],
    candidate_stops_by_day: dict[str, dict[str, list[dict[str, Any]]]],
    outliers: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], set[str], dict[tuple[str, str], set[str]]]:
    visits_by_block = group_visits_by_block(visits)
    assignments_by_day_device: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    unassigned_blocks: list[dict[str, Any]] = []
    for assignment in assignments:
        device = assignment["dispositivo_asignado"]
        if not device:
            unassigned_blocks.append(assignment)
            continue
        assignments_by_day_device[(assignment["fecha"], device)].append(assignment)

    matches: list[dict[str, Any]] = []
    matched_stop_ids: set[str] = set()
    matched_stops_by_day_device: dict[tuple[str, str], set[str]] = defaultdict(set)

    for (date_key, device), daily_assignments in assignments_by_day_device.items():
        stops = candidate_stops_by_day.get(date_key, {}).get(device, [])
        last_stop_index = -1
        for assignment in sorted(
            daily_assignments, key=lambda item: item["block_sequence"]
        ):
            block_id = assignment["bloque_diario"]
            for visit in visits_by_block[block_id]:
                if visit["lat"] is None or visit["lon"] is None:
                    matches.append(
                        {
                            "fecha": visit["fecha"],
                            "dispositivo": device,
                            "visita_programada_id": visit["visita_programada_id"],
                            "paciente_id": visit["paciente_id"],
                            "bloque_diario": block_id,
                            "stop_id": "",
                            "scheduled_time": dt_to_iso(visit["horario_programado"]),
                            "distance_m": None,
                            "delta_min": None,
                            "confianza_match": "sin_match",
                            "observacion": "Direccion sin geocodificacion utilizable.",
                        }
                    )
                    outliers.append(
                        {
                            "category": "visit_unmatched",
                            "fecha": visit["fecha"],
                            "dispositivo": device,
                            "bloque_diario": block_id,
                            "visita_programada_id": visit["visita_programada_id"],
                            "reference": visit["paciente_id"],
                            "detail": "Direccion sin geocodificacion; visita no comparable espacialmente.",
                        }
                    )
                    continue

                best_index = None
                best_payload = None
                for index, stop in enumerate(stops):
                    if index <= last_stop_index:
                        continue
                    if stop["stop_id"] in matched_stop_ids:
                        continue
                    delta_min = abs((stop["inicio"] - visit["horario_programado"]).total_seconds()) / 60
                    if delta_min > 150:
                        if stop["inicio"] > visit["horario_programado"] + timedelta(minutes=150):
                            break
                        continue
                    distance_m = haversine_m(visit["lat"], visit["lon"], stop["lat"], stop["lon"])
                    if distance_m > 1500:
                        continue
                    cost = distance_m / 500 + delta_min / 30
                    if best_payload is None or cost < best_payload["cost"]:
                        best_index = index
                        best_payload = {
                            "cost": cost,
                            "stop": stop,
                            "delta_min": delta_min,
                            "distance_m": distance_m,
                        }

                if best_payload and best_index is not None:
                    stop = best_payload["stop"]
                    distance_m = best_payload["distance_m"]
                    delta_min = best_payload["delta_min"]
                    matched_stop_ids.add(stop["stop_id"])
                    matched_stops_by_day_device[(date_key, device)].add(stop["stop_id"])
                    last_stop_index = best_index
                    confidence = (
                        "alta"
                        if distance_m <= 500 and delta_min <= 60
                        else "media"
                    )
                    matches.append(
                        {
                            "fecha": visit["fecha"],
                            "dispositivo": device,
                            "visita_programada_id": visit["visita_programada_id"],
                            "paciente_id": visit["paciente_id"],
                            "bloque_diario": block_id,
                            "stop_id": stop["stop_id"],
                            "scheduled_time": dt_to_iso(visit["horario_programado"]),
                            "distance_m": round(distance_m, 1),
                            "delta_min": round(delta_min, 1),
                            "confianza_match": confidence,
                            "observacion": "",
                        }
                    )
                else:
                    matches.append(
                        {
                            "fecha": visit["fecha"],
                            "dispositivo": device,
                            "visita_programada_id": visit["visita_programada_id"],
                            "paciente_id": visit["paciente_id"],
                            "bloque_diario": block_id,
                            "stop_id": "",
                            "scheduled_time": dt_to_iso(visit["horario_programado"]),
                            "distance_m": None,
                            "delta_min": None,
                            "confianza_match": "sin_match",
                            "observacion": "Sin parada compatible dentro de umbral temporal/espacial.",
                        }
                    )
                    outliers.append(
                        {
                            "category": "visit_unmatched",
                            "fecha": visit["fecha"],
                            "dispositivo": device,
                            "bloque_diario": block_id,
                            "visita_programada_id": visit["visita_programada_id"],
                            "reference": visit["paciente_id"],
                            "detail": "Sin parada GPS dentro de <=1500 m y <=150 min.",
                        }
                    )

    for assignment in unassigned_blocks:
        block_id = assignment["bloque_diario"]
        for visit in visits_by_block.get(block_id, []):
            matches.append(
                {
                    "fecha": visit["fecha"],
                    "dispositivo": "",
                    "visita_programada_id": visit["visita_programada_id"],
                    "paciente_id": visit["paciente_id"],
                    "bloque_diario": block_id,
                    "stop_id": "",
                    "scheduled_time": dt_to_iso(visit["horario_programado"]),
                    "distance_m": None,
                    "delta_min": None,
                    "confianza_match": "sin_match",
                    "observacion": (
                        "Bloque sin dispositivo asignable con evidencia suficiente. "
                        + assignment["motivo_ambiguedad"]
                    ).strip(),
                }
            )

    return matches, matched_stop_ids, matched_stops_by_day_device


def build_daily_vehicle_summary(
    events: list[dict[str, Any]],
    assignments: list[dict[str, Any]],
    visits: list[dict[str, Any]],
    matches: list[dict[str, Any]],
    matched_stops_by_day_device: dict[tuple[str, str], set[str]],
    outliers: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    visits_by_block = group_visits_by_block(visits)
    blocks_by_day_device: dict[tuple[str, str], list[str]] = defaultdict(list)
    for assignment in assignments:
        device = assignment["dispositivo_asignado"]
        if device:
            blocks_by_day_device[(assignment["fecha"], device)].append(assignment["bloque_diario"])

    matches_by_day_device: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for match in matches:
        matches_by_day_device[(match["fecha"], match["dispositivo"])].append(match)

    events_by_day_device: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        events_by_day_device[(event["date"], event["device"])].append(event)

    summary_rows: list[dict[str, Any]] = []
    for (date_key, device), daily_events in sorted(events_by_day_device.items()):
        assigned_blocks = blocks_by_day_device.get((date_key, device), [])
        programmed_visits = sum(len(visits_by_block[block_id]) for block_id in assigned_blocks)
        device_matches = matches_by_day_device.get((date_key, device), [])
        high = sum(1 for match in device_matches if match["confianza_match"] == "alta")
        medium = sum(1 for match in device_matches if match["confianza_match"] == "media")
        unmatched = sum(1 for match in device_matches if match["confianza_match"] == "sin_match")
        movement_km = sum(event["distancia_km"] or 0 for event in daily_events if event["tipo_evento"] == "Movimiento")
        movement_minutes = sum(event["duracion_seg"] for event in daily_events if event["tipo_evento"] == "Movimiento") / 60
        stopped_minutes = sum(event["duracion_seg"] for event in daily_events if event["tipo_evento"] == "Detenido") / 60
        first_departure = next(
            (event["inicio"] for event in daily_events if not event["base_stop"]),
            None,
        )
        last_return = None
        for event in reversed(daily_events):
            if not event["base_stop"]:
                last_return = event["fin"]
                break

        matched_stop_ids = matched_stops_by_day_device.get((date_key, device), set())
        extra_stops = 0
        extra_stop_minutes = 0.0
        for event in daily_events:
            if (
                event["tipo_evento"] == "Detenido"
                and not event["base_stop"]
                and event["duracion_seg"] > 600
                and event["stop_id"] not in matched_stop_ids
            ):
                extra_stops += 1
                extra_stop_minutes += event["duracion_seg"] / 60
                outliers.append(
                    {
                        "category": "extra_stop",
                        "fecha": date_key,
                        "dispositivo": device,
                        "bloque_diario": "",
                        "visita_programada_id": "",
                        "reference": event["stop_id"],
                        "detail": (
                            f"Parada >10 min no asociada a visita ni base "
                            f"({round(event['duracion_seg'] / 60, 1)} min)."
                        ),
                    }
                )

        matched_total = high + medium
        summary_rows.append(
            {
                "fecha": date_key,
                "dispositivo": device,
                "visitas_programadas": programmed_visits,
                "visitas_matcheadas_alta": high,
                "visitas_matcheadas_media": medium,
                "visitas_sin_match": unmatched,
                "km_totales": round(movement_km, 2),
                "tiempo_movimiento_min": round(movement_minutes, 1),
                "tiempo_detenido_min": round(stopped_minutes, 1),
                "paradas_extra": extra_stops,
                "tiempo_detenido_no_asistencial_min": round(extra_stop_minutes, 1),
                "salida_primera": dt_to_iso(first_departure),
                "llegada_ultima": dt_to_iso(last_return),
                "cumplimiento_pct": round((matched_total / programmed_visits) * 100, 1) if programmed_visits else None,
                "km_por_visita_programada": round(movement_km / programmed_visits, 2) if programmed_visits else None,
            }
        )

    return summary_rows


def build_outlier_summary(
    outliers: list[dict[str, Any]], assignments: list[dict[str, Any]]
) -> None:
    for assignment in assignments:
        if assignment["confianza_asignacion"] == "ambiguo":
            outliers.append(
                {
                    "category": "block_assignment_ambiguous",
                    "fecha": assignment["fecha"],
                    "dispositivo": assignment["dispositivo_asignado"],
                    "bloque_diario": assignment["bloque_diario"],
                    "visita_programada_id": "",
                    "reference": assignment["lider_bloque"],
                    "detail": assignment["motivo_ambiguedad"],
                }
            )


def format_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, float):
        return round(value, 3)
    return value


def write_table(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT
    for row in rows:
        ws.append([format_value(row.get(header, "")) for header in headers])
    for column_cells in ws.columns:
        max_length = max(len(clean_text(cell.value)) for cell in column_cells[:200]) if column_cells else 10
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 32)
    ws.freeze_panes = "A2"


def write_workbook(
    visits: list[dict[str, Any]],
    events: list[dict[str, Any]],
    geocodes: dict[str, GeocodeResult],
    assignments: list[dict[str, Any]],
    matches: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
    outliers: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    planned_headers = [
        "visita_programada_id",
        "fecha",
        "hoja",
        "archivo_fuente",
        "fila_fuente",
        "bloque_diario",
        "block_sequence",
        "lider_bloque",
        "hora_programada",
        "medico",
        "fono",
        "kine",
        "enfermera",
        "tens",
        "paciente",
        "paciente_id",
        "tipo_visita",
        "direccion",
        "direccion_normalizada",
        "telefono",
        "lat",
        "lon",
        "geocode_quality",
    ]
    gps_headers = [
        "event_id",
        "stop_id",
        "device",
        "date",
        "tipo_evento",
        "inicio",
        "fin",
        "duracion_seg",
        "ralenti_seg",
        "conductor",
        "lat",
        "lon",
        "distancia_km",
        "vel_max_kph",
        "vel_media_kph",
        "candidate_stop",
        "base_stop",
    ]
    geocode_headers = [
        "direccion_normalizada",
        "lat",
        "lon",
        "fuente_geocoder",
        "calidad_geocoder",
        "cache_key",
        "consulta_usada",
        "display_name",
        "token_overlap",
    ]
    assignment_headers = [
        "fecha",
        "bloque_diario",
        "block_sequence",
        "lider_bloque",
        "dispositivo_asignado",
        "score_asignacion",
        "confianza_asignacion",
        "motivo_ambiguedad",
        "second_best_score",
        "candidate_count",
        "high_candidates",
        "medium_candidates",
        "impossible_overlaps",
    ]
    match_headers = [
        "fecha",
        "dispositivo",
        "visita_programada_id",
        "paciente_id",
        "bloque_diario",
        "stop_id",
        "scheduled_time",
        "distance_m",
        "delta_min",
        "confianza_match",
        "observacion",
    ]
    summary_headers = [
        "fecha",
        "dispositivo",
        "visitas_programadas",
        "visitas_matcheadas_alta",
        "visitas_matcheadas_media",
        "visitas_sin_match",
        "km_totales",
        "tiempo_movimiento_min",
        "tiempo_detenido_min",
        "paradas_extra",
        "tiempo_detenido_no_asistencial_min",
        "salida_primera",
        "llegada_ultima",
        "cumplimiento_pct",
        "km_por_visita_programada",
    ]
    outlier_headers = [
        "category",
        "fecha",
        "dispositivo",
        "bloque_diario",
        "visita_programada_id",
        "reference",
        "detail",
    ]

    write_table(workbook.create_sheet("planned_visits"), visits, planned_headers)
    write_table(workbook.create_sheet("gps_events"), events, gps_headers)
    write_table(
        workbook.create_sheet("geocoded_addresses"),
        [geocodes[key].as_row() for key in sorted(geocodes)],
        geocode_headers,
    )
    write_table(workbook.create_sheet("daily_block_assignment"), assignments, assignment_headers)
    write_table(workbook.create_sheet("visit_matches"), matches, match_headers)
    write_table(workbook.create_sheet("daily_vehicle_summary"), summary_rows, summary_headers)
    write_table(workbook.create_sheet("outliers"), outliers, outlier_headers)

    workbook.save(WORKBOOK_PATH)


def build_report(
    visits: list[dict[str, Any]],
    events: list[dict[str, Any]],
    geocodes: dict[str, GeocodeResult],
    assignments: list[dict[str, Any]],
    matches: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
    outliers: list[dict[str, Any]],
) -> str:
    total_visits = len(visits)
    match_counter = Counter(match["confianza_match"] for match in matches)
    high_matches = match_counter["alta"]
    medium_matches = match_counter["media"]
    unmatched = match_counter["sin_match"]
    matched_total = high_matches + medium_matches
    geocoded_total = sum(1 for result in geocodes.values() if result.lat is not None)
    geocode_rate = (geocoded_total / len(geocodes) * 100) if geocodes else 0.0
    ambiguous_blocks = sum(
        1 for assignment in assignments if assignment["confianza_asignacion"] == "ambiguo"
    )
    reserve_blocks = sum(
        1 for assignment in assignments if assignment["confianza_asignacion"] == "asignado con reserva"
    )

    per_device: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in summary_rows:
        bucket = per_device[row["dispositivo"]]
        bucket["visitas_programadas"] += row["visitas_programadas"] or 0
        bucket["visitas_matcheadas"] += (row["visitas_matcheadas_alta"] or 0) + (row["visitas_matcheadas_media"] or 0)
        bucket["km_totales"] += row["km_totales"] or 0
        bucket["paradas_extra"] += row["paradas_extra"] or 0
        bucket["tiempo_detenido_no_asistencial_min"] += row["tiempo_detenido_no_asistencial_min"] or 0

    device_lines = []
    for device in sorted(per_device):
        stats = per_device[device]
        match_rate = (
            (stats["visitas_matcheadas"] / stats["visitas_programadas"]) * 100
            if stats["visitas_programadas"]
            else 0
        )
        device_lines.append(
            f"- `{device}`: {int(stats['visitas_programadas'])} visitas programadas, "
            f"{round(match_rate, 1)}% con match operativo, {round(stats['km_totales'], 1)} km, "
            f"{int(stats['paradas_extra'])} paradas extra."
        )

    worst_days = sorted(
        [row for row in summary_rows if row["visitas_programadas"]],
        key=lambda item: (
            item["cumplimiento_pct"] if item["cumplimiento_pct"] is not None else -1,
            -item["paradas_extra"],
        ),
    )[:5]
    worst_day_lines = [
        f"- `{row['fecha']}` / `{row['dispositivo']}`: cumplimiento {row['cumplimiento_pct']}%, "
        f"{row['visitas_sin_match']} sin match, {row['paradas_extra']} paradas extra."
        for row in worst_days
    ]

    locality_deviation: dict[str, list[float]] = defaultdict(list)
    visits_by_id = {visit["visita_programada_id"]: visit for visit in visits}
    for match in matches:
        if match["distance_m"] is None:
            continue
        visit = visits_by_id[match["visita_programada_id"]]
        locality = detect_locality(visit["direccion_normalizada"])
        locality_deviation[locality].append(float(match["distance_m"]))

    locality_lines = []
    for locality, distances in sorted(
        locality_deviation.items(), key=lambda item: mean(item[1]), reverse=True
    )[:5]:
        locality_lines.append(
            f"- `{locality}`: desviacion media {round(mean(distances), 1)} m en {len(distances)} visitas comparables."
        )

    recommendations = [
        "- Consolidar una llave diaria explicita `vehiculo/bloque` en la planilla para reducir las asignaciones ambiguas.",
        "- Estandarizar direcciones operativas antes de publicar la ruta diaria; las fallas de geocoding se concentran en domicilios con referencias abiertas.",
        "- Revisar dias con `paradas_extra` altas como proxy de tiempos muertos o desvio operativo, en especial cuando coinciden con bajo cumplimiento.",
        "- Mantener el 25 de marzo de 2026 fuera de comparaciones de ejecucion hasta contar con telemetria del mismo dia.",
    ]

    report = f"""# Análisis operativo de rutas HD Hospital de San Carlos

## 1. Resumen del período
- Ventana comparable: **2026-01-01** a **2026-03-24**.
- Dispositivos GPS comparables: **{len({event['device'] for event in events})}**.
- Visitas programadas comparables: **{total_visits}**.
- Direcciones únicas geocodificables: **{len(geocodes)}**; resueltas: **{geocoded_total}** ({round(geocode_rate, 1)}%).
- Matches operativos: **{matched_total}** ({round((matched_total / total_visits) * 100, 1)}%) entre `alta` y `media`.
- Visitas sin match: **{unmatched}**.
- Bloques ambiguos: **{ambiguous_blocks}**; bloques con reserva: **{reserve_blocks}**.

## 2. Perfil por vehículo
{chr(10).join(device_lines) if device_lines else "- Sin datos."}

## 3. Brechas de cumplimiento
{chr(10).join(worst_day_lines) if worst_day_lines else "- Sin brechas relevantes."}

## 4. Cuellos de botella operativos
{chr(10).join(locality_lines) if locality_lines else "- Sin desviaciones territoriales suficientes para ranking."}

## 5. Recomendaciones
{chr(10).join(recommendations)}

## Notas metodológicas
- El análisis es de **gestión operativa** y no reemplaza el juicio directivo ni clínico.
- Se anonimizaron pacientes en las tablas analíticas mediante identificadores `PAC-XXX`.
- La programación del **2026-03-25** queda fuera de la comparación principal por ausencia de telemetría utilizable en el CSV GPS.
- Los detalles fila a fila están en el workbook [`{WORKBOOK_PATH.name}`]({WORKBOOK_PATH.as_uri()}).
"""
    return report


def validate_counts(
    visits: list[dict[str, Any]],
    events: list[dict[str, Any]],
    geocodes: dict[str, GeocodeResult],
) -> None:
    devices = {event["device"] for event in events}
    if len(devices) != 3:
        raise RuntimeError(f"Se esperaban 3 dispositivos GPS y se obtuvieron {len(devices)}.")
    if len(visits) != 1573:
        raise RuntimeError(f"Se esperaban 1573 visitas comparables y se obtuvieron {len(visits)}.")
    max_event_date = max(datetime.fromisoformat(event["date"]) for event in events).date()
    if max_event_date != COMPARABLE_END:
        raise RuntimeError(
            f"La telemetria utilizable debia terminar en {COMPARABLE_END.isoformat()} y termino en {max_event_date}."
        )
    geocode_rate = sum(1 for result in geocodes.values() if result.lat is not None) / len(geocodes)
    if geocode_rate < 0.95:
        raise RuntimeError(
            f"La geocodificacion quedo bajo el umbral requerido: {round(geocode_rate * 100, 1)}%."
        )


def main() -> None:
    ensure_dirs()
    visits, outliers, block_meta = parse_planned_visits()
    events = parse_gps_events()
    base_clusters = build_base_clusters(events)
    print(f"Base clusters detectados: {json.dumps(base_clusters, ensure_ascii=False, indent=2)}")

    unique_addresses = sorted({visit["direccion_normalizada"] for visit in visits if visit["direccion_normalizada"]})
    geocodes = geocode_addresses(unique_addresses)
    enrich_visits_with_geocodes(visits, geocodes, outliers)
    validate_counts(visits, events, geocodes)

    block_visits = group_visits_by_block(visits)
    candidate_stops_by_day, events_by_day = build_daily_index(events)

    assignment_rows: list[dict[str, Any]] = []
    for date_key in sorted({visit["fecha"] for visit in visits}):
        assignment_rows.extend(
            assign_blocks_to_devices(
                date_key,
                block_visits,
                block_meta,
                candidate_stops_by_day,
                events_by_day,
            )
        )

    build_outlier_summary(outliers, assignment_rows)
    match_rows, _, matched_stops_by_day_device = match_assigned_visits(
        visits, assignment_rows, candidate_stops_by_day, outliers
    )
    summary_rows = build_daily_vehicle_summary(
        events, assignment_rows, visits, match_rows, matched_stops_by_day_device, outliers
    )

    assignment_rows.sort(key=lambda row: (row["fecha"], row["block_sequence"]))
    match_rows.sort(key=lambda row: (row["fecha"], row["dispositivo"], row["scheduled_time"]))
    summary_rows.sort(key=lambda row: (row["fecha"], row["dispositivo"]))
    outliers.sort(key=lambda row: (row["fecha"], row["category"], row["reference"]))

    write_workbook(visits, events, geocodes, assignment_rows, match_rows, summary_rows, outliers)
    report = build_report(visits, events, geocodes, assignment_rows, match_rows, summary_rows, outliers)
    REPORT_PATH.write_text(report, encoding="utf-8")

    print(f"Workbook generado: {WORKBOOK_PATH}")
    print(f"Reporte generado: {REPORT_PATH}")


if __name__ == "__main__":
    main()
