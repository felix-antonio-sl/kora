from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


MONTH_NAMES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


@dataclass(frozen=True)
class MonthSpec:
    year: int
    month: int

    @property
    def days(self) -> int:
        return monthrange(self.year, self.month)[1]

    @property
    def label(self) -> str:
        return f"{MONTH_NAMES[self.month]} {self.year}"


MONTHS = [
    MonthSpec(2024, 12),
    MonthSpec(2025, 1),
    MonthSpec(2025, 2),
    MonthSpec(2025, 3),
    MonthSpec(2025, 4),
    MonthSpec(2025, 5),
    MonthSpec(2025, 6),
    MonthSpec(2025, 7),
    MonthSpec(2025, 8),
    MonthSpec(2025, 9),
    MonthSpec(2025, 10),
    MonthSpec(2025, 11),
    MonthSpec(2025, 12),
    MonthSpec(2026, 1),
    MonthSpec(2026, 2),
]


THIN = Side(style="thin", color="D7DEE7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FILL = PatternFill("solid", fgColor="173F5F")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="EEF5FB")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
ALERT_FILL = PatternFill("solid", fgColor="FCE4D6")

TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(bold=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def lock_formula(cell) -> None:
    cell.protection = Protection(locked=True)
    cell.fill = FORMULA_FILL


def unlock_input(cell) -> None:
    cell.protection = Protection(locked=False)
    cell.fill = INPUT_FILL


def style_block(ws, cell_range: str, fill=None, font=None, alignment=None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment


def add_title(ws, title: str, end_col: int) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)


def build_portada(ws) -> None:
    ws.title = "00_PORTADA"
    add_title(ws, "HODOM - LIBRO OPERATIVO IDEAL", 6)

    lines = [
        "Para que sirve",
        "Base nueva para reemplazar la planilla antigua sin heredar su desorden ni sus errores.",
        "",
        "Que trae",
        "1. Parametros generales.",
        "2. Diccionario de indicadores.",
        "3. Tracker de migracion.",
        "4. Resumen global.",
        "5. Una hoja mensual estandar por cada mes.",
        "",
        "Como esta pensado",
        "Las celdas amarillas son para cargar datos.",
        "Las verdes calculan solas.",
        "Cada hoja mensual trae control diario y resumen mensual.",
        "Si no existe desglose por tipo, se puede cargar total informado de ingresos y/o egresos.",
        "",
        "Regla base",
        "No escribir encima de formulas. Si algo no cuadra, se corrige en la fila diaria, no en el resumen.",
    ]
    for idx, value in enumerate(lines, start=3):
        ws[f"A{idx}"] = value
    for col, width in {"A": 42, "B": 24, "C": 20, "D": 20, "E": 20, "F": 20}.items():
        ws.column_dimensions[col].width = width


def build_parametros(ws) -> None:
    ws.title = "01_PARAMETROS"
    add_title(ws, "PARAMETROS GENERALES", 4)

    rows = [
        ("Servicio", "HODOM"),
        ("Codigo", ""),
        ("Camas base por dia", 16),
        ("Responsable", ""),
        ("Centro / unidad", ""),
        ("Notas generales", ""),
    ]
    for row_idx, (label, value) in enumerate(rows, start=3):
        ws[f"A{row_idx}"] = label
        ws[f"A{row_idx}"].font = HEADER_FONT
        ws[f"A{row_idx}"].fill = SECTION_FILL
        ws[f"A{row_idx}"].border = BORDER
        ws[f"B{row_idx}"] = value
        ws[f"B{row_idx}"].border = BORDER
        unlock_input(ws[f"B{row_idx}"])

    ws["D3"] = "Criterios fijos"
    ws["D3"].font = HEADER_FONT
    ws["D3"].fill = SECTION_FILL
    ws["D3"].border = BORDER
    ws["D4"] = "Pacientes atendidos = censo inicial + ingresos"
    ws["D5"] = "Indice ocupacional = dias cama ocupadas / dias cama disponibles"
    ws["D6"] = "Promedio dia de estada = dias estada total / total egresos"
    ws["D7"] = "Si no hay desglose, se puede usar total informado del dia"
    for cell in ("D4", "D5", "D6", "D7"):
        ws[cell].border = BORDER

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 52


def build_diccionario(ws) -> None:
    ws.title = "02_DICCIONARIO"
    add_title(ws, "DICCIONARIO DE CAMPOS E INDICADORES", 4)

    headers = ["Elemento", "Tipo", "Que significa", "Como se obtiene"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(3, idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    rows = [
        ("Existencia dia anterior", "Dato / formula", "Pacientes con que parte el dia", "Dia 1 manual, resto arrastra desde camas ocupadas del dia previo"),
        ("Total ingresos", "Formula", "Todos los ingresos del dia", "Si hay desglose usa la suma; si no, usa el total informado"),
        ("Total ingresos informado", "Dato opcional", "Total del dia cuando no existe desglose", "Se carga solo si no hay detalle por tipo"),
        ("Total egresos", "Formula", "Todos los egresos del dia", "Si hay desglose usa la suma; si no, usa el total informado"),
        ("Total egresos informado", "Dato opcional", "Total del dia cuando no existe desglose", "Se carga solo si no hay detalle por tipo"),
        ("Dias cama ocupadas", "Formula", "Censo final del dia", "Existencia inicial + ingresos - egresos"),
        ("Pacientes atendidos del mes", "Formula", "Personas que pasaron por el servicio en el mes", "Censo inicial + ingresos totales"),
        ("Dias personas atendidas", "Formula", "Acumulado de estada del mes", "Suma de dias estada total"),
        ("Indice ocupacional", "Formula", "Nivel de ocupacion del mes", "Dias cama ocupadas / dias cama disponibles"),
        ("Promedio dia de estada", "Formula", "Promedio de permanencia", "Dias estada total / total egresos"),
        ("Balance del mes", "Control", "Chequea consistencia del mes", "Censo inicial + ingresos - egresos - censo final"),
    ]
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = BORDER
            if col_idx == 1:
                cell.fill = SECTION_FILL
                cell.font = HEADER_FONT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 58


def build_migracion(ws) -> None:
    ws.title = "03_MIGRACION"
    add_title(ws, "TRACKER DE MIGRACION", 6)

    headers = ["Mes", "Estado", "Fuente original", "Migrado por", "Validado", "Notas"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(3, idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    status_validation = DataValidation(
        type="list",
        formula1='"Pendiente,En carga,Validar,Listo"',
        allow_blank=True,
    )
    ws.add_data_validation(status_validation)

    for row_idx, spec in enumerate(MONTHS, start=4):
        ws[f"A{row_idx}"] = spec.label
        ws[f"A{row_idx}"].border = BORDER
        ws[f"B{row_idx}"] = "Pendiente"
        ws[f"B{row_idx}"].border = BORDER
        ws[f"C{row_idx}"].border = BORDER
        ws[f"D{row_idx}"].border = BORDER
        ws[f"E{row_idx}"].border = BORDER
        ws[f"F{row_idx}"].border = BORDER
        for col in "BCDEF":
            unlock_input(ws[f"{col}{row_idx}"])
        status_validation.add(ws[f"B{row_idx}"])

    for col, width in {"A": 18, "B": 14, "C": 40, "D": 18, "E": 12, "F": 38}.items():
        ws.column_dimensions[col].width = width


def build_month_sheet(ws, spec: MonthSpec) -> None:
    ws.title = spec.label
    ws.freeze_panes = "A7"
    add_title(ws, f"HODOM - {spec.label}", 26)

    meta = [
        ("Servicio", "=01_PARAMETROS!B3"),
        ("Codigo", "=01_PARAMETROS!B4"),
        ("Mes", MONTH_NAMES[spec.month]),
        ("Ano", spec.year),
        ("Camas base", "=01_PARAMETROS!B5"),
        ("Dias del mes", spec.days),
    ]
    positions = ["A2", "C2", "E2", "G2", "I2", "K2"]
    for pos, (label, value) in zip(positions, meta):
        col = pos[0]
        row = pos[1:]
        value_col = get_column_letter(ord(col) - 64 + 1)
        ws[f"{col}{row}"] = label
        ws[f"{col}{row}"].font = HEADER_FONT
        ws[f"{col}{row}"].fill = SECTION_FILL
        ws[f"{col}{row}"].border = BORDER
        ws[f"{value_col}{row}"] = value
        ws[f"{value_col}{row}"].border = BORDER
        if isinstance(value, str) and value.startswith("="):
            lock_formula(ws[f"{value_col}{row}"])
        else:
            unlock_input(ws[f"{value_col}{row}"])

    headers = [
        "Fecha",
        "Dia",
        "Existencia inicial",
        "Ing. urgencia",
        "Ing. hospitalizacion",
        "Ing. ambulatorio",
        "Ing. ley urgencia",
        "Ing. UGCC",
        "Ing. traslado",
        "Total ingresos",
        "Total ingresos informado",
        "Alta hogar / otro hosp.",
        "Traslado a otro servicio",
        "Fallecidos",
        "Total egresos",
        "Total egresos informado",
        "Ingr.-egr. mismo dia",
        "Camas disponibles",
        "Camas ocupadas",
        "Dias estada total",
        "Dias estada benef.",
        "Observaciones",
        "Estado dia",
        "Check camas",
        "Check balance",
        "Check faltantes",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(6, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = WRAP_CENTER

    first_row = 7
    last_row = first_row + spec.days - 1

    non_negative = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    ws.add_data_validation(non_negative)

    for day in range(1, spec.days + 1):
        row = first_row + day - 1
        ws[f"A{row}"] = f"=DATE({spec.year},{spec.month},{day})"
        ws[f"B{row}"] = day
        lock_formula(ws[f"A{row}"])
        lock_formula(ws[f"B{row}"])

        if day == 1:
            unlock_input(ws[f"C{row}"])
            non_negative.add(ws[f"C{row}"])
        else:
            ws[f"C{row}"] = f"=Q{row-1}"
            lock_formula(ws[f"C{row}"])

        for col in "DEFGHIKLNPQRTUV":
            unlock_input(ws[f"{col}{row}"])
            non_negative.add(ws[f"{col}{row}"])

        ws[f"J{row}"] = f'=IF(COUNTA(D{row}:I{row})>0,SUM(D{row}:I{row}),IF(K{row}="", "", K{row}))'
        ws[f"O{row}"] = f'=IF(COUNTA(L{row}:N{row})>0,SUM(L{row}:N{row}),IF(P{row}="", "", P{row}))'
        ws[f"S{row}"] = f'=IF(OR(C{row}="",J{row}="",O{row}=""),"",C{row}+J{row}-O{row})'
        ws[f"W{row}"] = (
            f'=IF(Z{row}=1,"FALTAN DATOS",'
            f'IF(S{row}<0,"OCUPACION NEGATIVA",'
            f'IF(X{row}=1,"SOBREOCUPACION",'
            f'IF(Y{row}<>0,"BALANCE RARO","OK"))))'
        )
        ws[f"X{row}"] = f'=IF(AND(R{row}<>"",S{row}>R{row}),1,0)'
        ws[f"Y{row}"] = f'=IF(OR(C{row}="",J{row}="",O{row}="",S{row}=""),"",C{row}+J{row}-O{row}-S{row})'
        ws[f"Z{row}"] = f'=IF(OR(R{row}="",A{row}="",B{row}="",C{row}="",J{row}="",O{row}=""),1,0)'

        for col in "JOSWXYZ":
            lock_formula(ws[f"{col}{row}"])

        for col_idx in range(1, 27):
            ws.cell(row, col_idx).border = BORDER

    total_row = last_row + 1
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = HEADER_FONT
    ws[f"A{total_row}"].fill = SECTION_FILL
    ws[f"A{total_row}"].border = BORDER
    for col in "DEFGHIJKLMNOPQRSTUV":
        ws[f"{col}{total_row}"] = f"=SUM({col}{first_row}:{col}{last_row})"
        ws[f"{col}{total_row}"].font = HEADER_FONT
        ws[f"{col}{total_row}"].border = BORDER
        lock_formula(ws[f"{col}{total_row}"])

    summary_row = total_row + 3
    ws[f"A{summary_row}"] = "RESUMEN DEL MES"
    ws[f"A{summary_row}"].font = HEADER_FONT
    ws[f"A{summary_row}"].fill = SECTION_FILL
    ws[f"A{summary_row}"].border = BORDER
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)

    summary = [
        ("Total ingresos", f"=J{total_row}"),
        ("Total egresos", f"=O{total_row}"),
        ("Altas hogar / otro hosp.", f"=L{total_row}"),
        ("Traslados a otro servicio", f"=M{total_row}"),
        ("Fallecidos", f"=N{total_row}"),
        ("Pacientes atendidos del mes", f"=C{first_row}+J{total_row}"),
        ("Dias personas atendidas", f"=T{total_row}"),
        ("Promedio camas", f"=AVERAGE(R{first_row}:R{last_row})"),
        ("Indice ocupacional", f'=IFERROR(S{total_row}/R{total_row},"")'),
        ("Promedio dia de estada", f'=IFERROR(T{total_row}/O{total_row},"")'),
        ("Censo final", f"=S{last_row}"),
        ("Balance del mes", f"=C{first_row}+J{total_row}-O{total_row}-S{last_row}"),
        ("Dias con sobreocupacion", f'=COUNTIF(W{first_row}:W{last_row},"SOBREOCUPACION")'),
        ("Dias con alerta", f'=COUNTIF(W{first_row}:W{last_row},"<>OK")'),
    ]
    for idx, (label, formula) in enumerate(summary, start=1):
        row = summary_row + idx
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = HEADER_FONT
        ws[f"A{row}"].fill = HEADER_FILL
        ws[f"A{row}"].border = BORDER
        ws[f"B{row}"] = formula
        ws[f"B{row}"].border = BORDER
        lock_formula(ws[f"B{row}"])
        if label == "Indice ocupacional":
            ws[f"B{row}"].number_format = "0.00%"
        elif "Promedio" in label:
            ws[f"B{row}"].number_format = "0.00"

    control_row = total_row + 3
    ws[f"F{control_row}"] = "CONTROL DEL MES"
    ws[f"F{control_row}"].font = HEADER_FONT
    ws[f"F{control_row}"].fill = SECTION_FILL
    ws[f"F{control_row}"].border = BORDER
    ws.merge_cells(start_row=control_row, start_column=6, end_row=control_row, end_column=9)

    controls = [
        ("Balance debe ser 0", f'=IF(B{summary_row+12}=0,"OK","REVISAR")'),
        ("No debe haber ocupacion negativa", f'=IF(COUNTIF(W{first_row}:W{last_row},"OCUPACION NEGATIVA")=0,"OK","REVISAR")'),
        ("No deberia faltar existencia inicial", f'=IF(C{first_row}="","REVISAR","OK")'),
        ("Dias con alertas", f"=B{summary_row+14}"),
    ]
    for idx, (label, formula) in enumerate(controls, start=1):
        row = control_row + idx
        ws[f"F{row}"] = label
        ws[f"F{row}"].font = HEADER_FONT
        ws[f"F{row}"].fill = HEADER_FILL
        ws[f"F{row}"].border = BORDER
        ws[f"G{row}"] = formula
        ws[f"G{row}"].border = BORDER
        lock_formula(ws[f"G{row}"])

    widths = {
        "A": 12,
        "B": 6,
        "C": 18,
        "D": 12,
        "E": 18,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 12,
        "J": 13,
        "K": 18,
        "L": 20,
        "M": 21,
        "N": 12,
        "O": 13,
        "P": 18,
        "Q": 16,
        "R": 16,
        "S": 15,
        "T": 16,
        "U": 16,
        "V": 24,
        "W": 18,
        "X": 12,
        "Y": 12,
        "Z": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.column_dimensions["X"].hidden = True
    ws.column_dimensions["Y"].hidden = True
    ws.column_dimensions["Z"].hidden = True

    ws.sheet_view.showGridLines = True
    ws.protection.sheet = True
    ws.protection.password = "hodom"
    ws.protection.enable()
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def build_resumen_global(ws) -> None:
    ws.title = "90_RESUMEN_GLOBAL"
    add_title(ws, "RESUMEN GLOBAL", 14)

    headers = [
        "Mes",
        "Ingresos",
        "Egresos",
        "Altas",
        "Traslados",
        "Fallecidos",
        "Pacientes atendidos",
        "Dias personas atendidas",
        "Promedio camas",
        "Indice ocupacional",
        "Promedio estada",
        "Censo final",
        "Dias con sobreocupacion",
        "Dias con alerta",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(3, idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for row_idx, spec in enumerate(MONTHS, start=4):
        name = f"'{spec.label}'"
        summary_start = spec.days + 11
        ws[f"A{row_idx}"] = spec.label
        ws[f"B{row_idx}"] = f"={name}!B{summary_start}"
        ws[f"C{row_idx}"] = f"={name}!B{summary_start+1}"
        ws[f"D{row_idx}"] = f"={name}!B{summary_start+2}"
        ws[f"E{row_idx}"] = f"={name}!B{summary_start+3}"
        ws[f"F{row_idx}"] = f"={name}!B{summary_start+4}"
        ws[f"G{row_idx}"] = f"={name}!B{summary_start+5}"
        ws[f"H{row_idx}"] = f"={name}!B{summary_start+6}"
        ws[f"I{row_idx}"] = f"={name}!B{summary_start+7}"
        ws[f"J{row_idx}"] = f"={name}!B{summary_start+8}"
        ws[f"K{row_idx}"] = f"={name}!B{summary_start+9}"
        ws[f"L{row_idx}"] = f"={name}!B{summary_start+10}"
        ws[f"M{row_idx}"] = f"={name}!B{summary_start+12}"
        ws[f"N{row_idx}"] = f"={name}!B{summary_start+13}"
        for col_idx in range(1, 15):
            cell = ws.cell(row_idx, col_idx)
            cell.border = BORDER
            if col_idx > 1:
                lock_formula(cell)
        ws[f"J{row_idx}"].number_format = "0.00%"
        ws[f"I{row_idx}"].number_format = "0.00"
        ws[f"K{row_idx}"].number_format = "0.00"

    for col, width in {
        "A": 18,
        "B": 10,
        "C": 10,
        "D": 10,
        "E": 12,
        "F": 10,
        "G": 18,
        "H": 20,
        "I": 14,
        "J": 16,
        "K": 16,
        "L": 10,
        "M": 18,
        "N": 14,
    }.items():
        ws.column_dimensions[col].width = width


def build_control_global(ws) -> None:
    ws.title = "99_CONTROL_GLOBAL"
    add_title(ws, "CONTROL GLOBAL", 6)

    headers = ["Mes", "Estado migracion", "Dias con alerta", "Dias con sobreocupacion", "Balance del mes", "Semaforo"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(3, idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for row_idx, spec in enumerate(MONTHS, start=4):
        name = f"'{spec.label}'"
        summary_start = spec.days + 11
        mig_row = row_idx
        ws[f"A{row_idx}"] = spec.label
        ws[f"B{row_idx}"] = f"=03_MIGRACION!B{mig_row}"
        ws[f"C{row_idx}"] = f"={name}!B{summary_start+13}"
        ws[f"D{row_idx}"] = f"={name}!B{summary_start+12}"
        ws[f"E{row_idx}"] = f"={name}!B{summary_start+11}"
        ws[f"F{row_idx}"] = f'=IF(OR(C{row_idx}>0,E{row_idx}<>0),"REVISAR","OK")'
        for col_idx in range(1, 7):
            cell = ws.cell(row_idx, col_idx)
            cell.border = BORDER
            if col_idx > 1:
                lock_formula(cell)

    for col, width in {"A": 18, "B": 16, "C": 14, "D": 18, "E": 14, "F": 12}.items():
        ws.column_dimensions[col].width = width


def create_workbook(output_path: Path) -> Path:
    wb = Workbook()
    build_portada(wb.active)
    build_parametros(wb.create_sheet("01_PARAMETROS"))
    build_diccionario(wb.create_sheet("02_DICCIONARIO"))
    build_migracion(wb.create_sheet("03_MIGRACION"))
    for spec in MONTHS:
        build_month_sheet(wb.create_sheet(spec.label), spec)
    build_resumen_global(wb.create_sheet("90_RESUMEN_GLOBAL"))
    build_control_global(wb.create_sheet("99_CONTROL_GLOBAL"))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    target = Path("tmp/spreadsheets/HODOM_ideal.xlsx")
    print(create_workbook(target))
